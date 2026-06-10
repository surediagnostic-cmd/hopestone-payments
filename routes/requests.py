import io
import json
from datetime import date, datetime, timezone
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file
from flask_login import login_required, current_user
from sqlalchemy import extract, func
from sqlalchemy.orm import subqueryload, joinedload

from extensions import db
from models import PaymentRequest, PaymentRequestItem, Branch, Category, Budget
from utils import format_naira

requests_bp = Blueprint("requests", __name__)


def _eager_pr():
    return PaymentRequest.query.options(
        subqueryload(PaymentRequest.items).joinedload(PaymentRequestItem.category),
        subqueryload(PaymentRequest.items).subqueryload(PaymentRequestItem.children).joinedload(PaymentRequestItem.category),
        joinedload(PaymentRequest.branch),
        joinedload(PaymentRequest.submitter),
    )


def _leaf_filter(q):
    """Filter a query on PaymentRequestItem to only return leaf items (no children).
    Use this on all financial aggregation queries to avoid double-counting parent roll-up rows."""
    parent_ids_sq = db.session.query(PaymentRequestItem.parent_id).filter(
        PaymentRequestItem.parent_id.isnot(None)
    ).subquery()
    return q.filter(~PaymentRequestItem.id.in_(parent_ids_sq))


# ── Dashboard ────────────────────────────────────────────────────────────────

@requests_bp.route("/dashboard")
@login_required
def dashboard():
    try:
        return _dashboard_inner()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        print(f"[dashboard error]: {e}", flush=True)
        return (
            f"<html><body style='font-family:sans-serif;padding:40px;"
            f"background:#f5f7fa;color:#2c3e50;'>"
            f"<h2>Dashboard temporarily unavailable.</h2>"
            f"<a href='/dashboard' style='color:#3a9fd8;'>↩ Retry</a>"
            f"<br><small style='color:#999'>Error: {str(e)}</small>"
            f"</body></html>"
        ), 200


def _dashboard_inner():
    now = datetime.now(timezone.utc)
    months_list = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    if current_user.is_mds:
        year_filter = int(request.args.get("year", now.year))
        month_filter = int(request.args.get("month", now.month))   # 0 = all months
        branch_id_filter = int(request.args.get("branch_id", 0))   # 0 = all branches

        all_branches = Branch.query.order_by(Branch.name).all()

        def _apply_period(q):
            q = q.filter(extract("year", PaymentRequest.created_at) == year_filter)
            if month_filter:
                q = q.filter(extract("month", PaymentRequest.created_at) == month_filter)
            if branch_id_filter:
                q = q.filter(PaymentRequest.branch_id == branch_id_filter)
            return q

        # KPI 1: all pending (global — always what needs attention now)
        all_pending_count = PaymentRequest.query.filter_by(status="pending").count()

        # KPI 2: approved amount in period
        approved_amt = _apply_period(
            db.session.query(func.sum(PaymentRequest.approved_amount))
            .filter(PaymentRequest.status == "approved")
        ).scalar() or 0

        # KPI 3: total pending amount (global)
        total_pending_amt = (
            db.session.query(func.sum(PaymentRequest.requested_amount))
            .filter(PaymentRequest.status == "pending")
            .scalar() or 0
        )

        # KPI 4: approval rate in period
        period_approved_count = _apply_period(
            PaymentRequest.query.filter(PaymentRequest.status == "approved")
        ).count()
        period_rejected_count = _apply_period(
            PaymentRequest.query.filter(PaymentRequest.status == "rejected")
        ).count()
        period_reviewed = period_approved_count + period_rejected_count
        approval_rate = round((period_approved_count / period_reviewed * 100) if period_reviewed else 0)

        # KPI 5 & 6: direct costs and overhead (approved leaf items in period)
        def _item_cost(cost_type):
            q = (
                db.session.query(func.sum(PaymentRequestItem.amount))
                .join(PaymentRequest, PaymentRequest.id == PaymentRequestItem.request_id)
                .join(Category, Category.id == PaymentRequestItem.category_id)
                .filter(
                    PaymentRequest.status == "approved",
                    Category.cost_type == cost_type,
                    extract("year", PaymentRequest.created_at) == year_filter,
                )
            )
            if month_filter:
                q = q.filter(extract("month", PaymentRequest.created_at) == month_filter)
            if branch_id_filter:
                q = q.filter(PaymentRequest.branch_id == branch_id_filter)
            return _leaf_filter(q).scalar() or 0

        direct_costs_amt = _item_cost("direct_cost")
        overhead_amt = _item_cost("overhead")

        # Status pill counts
        uploaded_count = _apply_period(
            PaymentRequest.query.filter(
                PaymentRequest.upload_status == "uploaded",
                PaymentRequest.status == "approved",
            )
        ).count()
        status_counts = {
            "pending": all_pending_count,
            "approved": period_approved_count,
            "rejected": period_rejected_count,
            "uploaded": uploaded_count,
        }

        # Bar chart: approved amount by branch in period
        bq = (
            db.session.query(Branch.name, func.sum(PaymentRequest.approved_amount))
            .join(PaymentRequest, PaymentRequest.branch_id == Branch.id)
            .filter(
                PaymentRequest.status == "approved",
                extract("year", PaymentRequest.created_at) == year_filter,
            )
        )
        if month_filter:
            bq = bq.filter(extract("month", PaymentRequest.created_at) == month_filter)
        if branch_id_filter:
            bq = bq.filter(PaymentRequest.branch_id == branch_id_filter)
        branch_rows = bq.group_by(Branch.name).order_by(func.sum(PaymentRequest.approved_amount).desc()).all()
        branch_labels = [r[0] for r in branch_rows]
        branch_values = [float(r[1] or 0) for r in branch_rows]

        # Donut chart + category table (leaf items only — no double-counting of parent roll-ups)
        cq = (
            db.session.query(
                Category.name,
                Category.cost_type,
                func.sum(PaymentRequestItem.amount).label("total"),
                func.count(PaymentRequestItem.id).label("cnt"),
            )
            .join(PaymentRequestItem, PaymentRequestItem.category_id == Category.id)
            .join(PaymentRequest, PaymentRequest.id == PaymentRequestItem.request_id)
            .filter(
                PaymentRequest.status == "approved",
                extract("year", PaymentRequest.created_at) == year_filter,
            )
        )
        if month_filter:
            cq = cq.filter(extract("month", PaymentRequest.created_at) == month_filter)
        if branch_id_filter:
            cq = cq.filter(PaymentRequest.branch_id == branch_id_filter)
        cq = _leaf_filter(cq)
        cat_rows = cq.group_by(Category.name, Category.cost_type).order_by(func.sum(PaymentRequestItem.amount).desc()).all()

        cat_total = sum(float(r.total or 0) for r in cat_rows)
        cat_data = [
            {
                "name": r.name,
                "cost_type": r.cost_type,
                "total": float(r.total or 0),
                "count": r.cnt,
                "share": round(float(r.total or 0) / cat_total * 100, 1) if cat_total else 0,
            }
            for r in cat_rows
        ]
        cat_labels = [d["name"] for d in cat_data]
        cat_values = [d["total"] for d in cat_data]

        # Recent transactions in period (limit 25)
        recent_txns = _apply_period(
            _eager_pr()
        ).order_by(PaymentRequest.created_at.desc()).limit(25).all()

        # Pending for review (global, for the pending section below charts)
        pending = _eager_pr().filter_by(status="pending").order_by(PaymentRequest.created_at.desc()).all()

        period_label = (months_list[month_filter - 1] + " " if month_filter else "") + str(year_filter)

        # ── Budget vs actual for dashboard card ───────────────────────────
        budget_month  = month_filter if month_filter else now.month
        budget_year   = year_filter
        budget_total  = float(
            db.session.query(func.sum(Budget.monthly_amount))
            .filter_by(year=budget_year, month=budget_month)
            .scalar() or 0
        )
        budget_pct    = round(float(approved_amt) / budget_total * 100, 1) if budget_total > 0 else None
        budget_label  = months_list[budget_month - 1] + " " + str(budget_year)

        return render_template(
            "dashboard.html",
            year_filter=year_filter,
            month_filter=month_filter,
            branch_id_filter=branch_id_filter,
            all_branches=all_branches,
            months_list=months_list,
            now=now,
            period_label=period_label,
            all_pending_count=all_pending_count,
            approved_amt=approved_amt,
            total_pending_amt=total_pending_amt,
            approval_rate=approval_rate,
            direct_costs_amt=direct_costs_amt,
            overhead_amt=overhead_amt,
            status_counts=status_counts,
            branch_labels=branch_labels,
            branch_values=branch_values,
            cat_labels=cat_labels,
            cat_values=cat_values,
            cat_data=cat_data,
            recent_txns=recent_txns,
            pending=pending,
            budget_total=budget_total,
            budget_pct=budget_pct,
            budget_label=budget_label,
            format_naira=format_naira,
        )
    else:
        acc_year  = int(request.args.get("year",  now.year))
        acc_month = int(request.args.get("month", now.month))  # 0 = all months

        def _my_period(q):
            q = q.filter(
                PaymentRequest.submitted_by_id == current_user.id,
                extract("year", PaymentRequest.created_at) == acc_year,
            )
            if acc_month:
                q = q.filter(extract("month", PaymentRequest.created_at) == acc_month)
            return q

        my_requests = _my_period(
            _eager_pr()
        ).order_by(PaymentRequest.created_at.desc()).limit(25).all()

        # Global pending — always show what needs MD attention regardless of period
        my_pending = PaymentRequest.query.filter_by(
            submitted_by_id=current_user.id, status="pending"
        ).count()

        my_approved_count = _my_period(
            PaymentRequest.query.filter(PaymentRequest.status == "approved")
        ).count()

        my_rejected_count = _my_period(
            PaymentRequest.query.filter(PaymentRequest.status == "rejected")
        ).count()

        my_total_approved = _my_period(
            db.session.query(func.sum(PaymentRequest.approved_amount))
            .filter(PaymentRequest.status == "approved")
        ).scalar() or 0

        my_total_requested = _my_period(
            db.session.query(func.sum(PaymentRequest.requested_amount))
        ).scalar() or 0

        acc_months_list = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        acc_period_label = (acc_months_list[acc_month - 1] + " " if acc_month else "") + str(acc_year)

        return render_template(
            "dashboard.html",
            my_requests=my_requests,
            my_pending=my_pending,
            my_approved_count=my_approved_count,
            my_rejected_count=my_rejected_count,
            my_total_approved=my_total_approved,
            my_total_requested=my_total_requested,
            acc_year=acc_year,
            acc_month=acc_month,
            acc_period_label=acc_period_label,
            acc_months_list=acc_months_list,
            now=now,
            format_naira=format_naira,
        )


# ── New Request ──────────────────────────────────────────────────────────────

@requests_bp.route("/requests/new", methods=["GET", "POST"])
@login_required
def new_request():
    if current_user.is_mds:
        branches = Branch.query.filter_by(is_active=True).all()
    else:
        branches = [b for b in current_user.branches if b.is_active]

    categories = Category.query.filter_by(is_active=True).order_by(Category.cost_type, Category.name).all()

    if request.method == "POST":
        try:
            payment_date = datetime.strptime(request.form["date"], "%Y-%m-%d").date()
            branch_id = int(request.form["branch_id"])
            beneficiary_name = request.form["beneficiary_name"].strip()
            beneficiary_account = request.form["beneficiary_account"].strip()
            beneficiary_bank = request.form["beneficiary_bank"].strip()
            bank_code = request.form.get("bank_code", "").strip() or None

            if not current_user.is_mds:
                branch_ids = [b.id for b in current_user.branches]
                if branch_id not in branch_ids:
                    flash("Invalid branch selection.", "danger")
                    return redirect(url_for("requests.new_request"))

            items_json_str = request.form.get("items_json", "[]")
            try:
                tree = json.loads(items_json_str)
            except (json.JSONDecodeError, ValueError):
                tree = []

            if not tree or not any(node.get("desc", "").strip() for node in tree):
                flash("At least one line item is required.", "danger")
                return render_template("new_request.html", branches=branches, categories=categories,
                                       today=date.today().isoformat())

            # Build the PaymentRequest first, then items
            pr = PaymentRequest(
                reference=PaymentRequest.generate_reference(),
                date=payment_date,
                branch_id=branch_id,
                beneficiary_name=beneficiary_name,
                beneficiary_account=beneficiary_account,
                beneficiary_bank=beneficiary_bank,
                bank_code=bank_code,
                requested_amount=Decimal("0"),  # updated below
                submitted_by_id=current_user.id,
            )
            db.session.add(pr)
            db.session.flush()  # get pr.id

            total = Decimal("0")
            for node in tree:
                desc = node.get("desc", "").strip()
                if not desc:
                    continue
                children_data = node.get("children", [])

                if children_data:
                    # Parent is a roll-up row — qty=1, rate=0, amount = sum of children
                    parent_item = PaymentRequestItem(
                        request_id=pr.id,
                        description=desc,
                        note=node.get("note", "").strip() or None,
                        category_id=int(node.get("cat_id", 1)),
                        quantity=1,
                        rate=Decimal("0"),
                        amount=Decimal("0"),  # updated after children saved
                    )
                    db.session.add(parent_item)
                    db.session.flush()  # get parent_item.id

                    child_total = Decimal("0")
                    for child in children_data:
                        cdesc = child.get("desc", "").strip()
                        if not cdesc:
                            continue
                        cqty = int(child.get("qty") or 1)
                        crate = Decimal(str(child.get("rate") or 0))
                        camount = cqty * crate
                        child_total += camount
                        db.session.add(PaymentRequestItem(
                            request_id=pr.id,
                            parent_id=parent_item.id,
                            description=cdesc,
                            note=child.get("note", "").strip() or None,
                            category_id=int(child.get("cat_id", 1)),
                            quantity=cqty,
                            rate=crate,
                            amount=camount,
                        ))
                    parent_item.amount = child_total
                    total += child_total
                else:
                    qty = int(node.get("qty") or 1)
                    rate = Decimal(str(node.get("rate") or 0))
                    amount = qty * rate
                    total += amount
                    db.session.add(PaymentRequestItem(
                        request_id=pr.id,
                        description=desc,
                        note=node.get("note", "").strip() or None,
                        category_id=int(node.get("cat_id", 1)),
                        quantity=qty,
                        rate=rate,
                        amount=amount,
                    ))

            pr.requested_amount = total
            db.session.commit()
            flash(f"Payment request {pr.reference} submitted successfully.", "success")
            return redirect(url_for("requests.view_request", req_id=pr.id))

        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            flash(f"Error submitting request: {str(e)}", "danger")

    return render_template("new_request.html", branches=branches, categories=categories,
                           today=date.today().isoformat())


# ── List Requests ────────────────────────────────────────────────────────────

@requests_bp.route("/requests/", strict_slashes=False)
@login_required
def list_requests():
    try:
        q = _eager_pr()
        if not current_user.is_mds:
            q = q.filter_by(submitted_by_id=current_user.id)

        status_filter = request.args.get("status", "")
        branch_filter = request.args.get("branch_id", "")
        month_filter = request.args.get("month", "")

        if status_filter:
            q = q.filter_by(status=status_filter)
        if branch_filter:
            q = q.filter_by(branch_id=int(branch_filter))
        if month_filter:
            try:
                year, month = month_filter.split("-")
                q = q.filter(
                    extract("year", PaymentRequest.created_at) == int(year),
                    extract("month", PaymentRequest.created_at) == int(month),
                )
            except Exception:
                pass

        all_requests = q.order_by(PaymentRequest.created_at.desc()).all()
        branches = Branch.query.filter_by(is_active=True).all()

        return render_template(
            "list_requests.html",
            requests=all_requests,
            branches=branches,
            format_naira=format_naira,
            status_filter=status_filter,
            branch_filter=branch_filter,
            month_filter=month_filter,
        )
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"Error loading requests: {str(e)}", "danger")
        return redirect(url_for("requests.dashboard"))


# ── View / Mark Uploaded / Delete ────────────────────────────────────────────

@requests_bp.route("/requests/<int:req_id>")
@login_required
def view_request(req_id):
    pr = _eager_pr().filter_by(id=req_id).first_or_404()
    if not current_user.is_mds and pr.submitted_by_id != current_user.id:
        flash("Access denied.", "danger")
        return redirect(url_for("requests.dashboard"))
    return render_template("request_detail.html", pr=pr, format_naira=format_naira)


@requests_bp.route("/requests/<int:req_id>/mark-uploaded", methods=["POST"])
@login_required
def mark_uploaded(req_id):
    pr = PaymentRequest.query.get_or_404(req_id)
    if not current_user.is_mds and pr.submitted_by_id != current_user.id:
        flash("Access denied.", "danger")
        return redirect(url_for("requests.dashboard"))
    if pr.status != "approved":
        flash("Only approved requests can be marked as uploaded.", "warning")
        return redirect(url_for("requests.view_request", req_id=req_id))
    pr.upload_status = "uploaded"
    db.session.commit()
    flash(f"{pr.reference} marked as uploaded to bank.", "success")
    return redirect(url_for("requests.view_request", req_id=req_id))


@requests_bp.route("/requests/<int:req_id>/delete", methods=["POST"])
@login_required
def delete_request(req_id):
    if not current_user.is_mds:
        flash("Access denied.", "danger")
        return redirect(url_for("requests.dashboard"))
    pr = PaymentRequest.query.get_or_404(req_id)
    ref = pr.reference
    db.session.delete(pr)
    db.session.commit()
    flash(f"{ref} has been deleted.", "success")
    return redirect(url_for("requests.list_requests"))


# ── Bulk Upload ───────────────────────────────────────────────────────────────

@requests_bp.route("/requests/bulk-upload", methods=["GET", "POST"])
@login_required
def bulk_upload():
    if request.method == "POST":
        if "file" not in request.files or not request.files["file"].filename:
            flash("Please select a file to upload.", "danger")
            return redirect(url_for("requests.bulk_upload"))

        file = request.files["file"]
        if not file.filename.endswith(".xlsx"):
            flash("Only .xlsx files are accepted.", "danger")
            return redirect(url_for("requests.bulk_upload"))

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            branches_map = {b.name.lower(): b for b in Branch.query.filter_by(is_active=True).all()}
            cats_map = {c.name.lower(): c for c in Category.query.filter_by(is_active=True).all()}

            errors = []
            created = 0

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                sample_check = str(row[2] or "").strip().lower()
                if sample_check in ("sample vendor", "example", ""):
                    continue

                try:
                    row_date = row[0]
                    if isinstance(row_date, str):
                        row_date = datetime.strptime(row_date.strip(), "%Y-%m-%d").date()
                    elif hasattr(row_date, "date"):
                        row_date = row_date.date()

                    branch_name = str(row[1] or "").strip().lower()
                    branch = branches_map.get(branch_name)
                    if not branch:
                        errors.append(f"Row {row_num}: Unknown branch '{row[1]}'")
                        continue

                    if not current_user.is_mds:
                        user_branch_ids = [b.id for b in current_user.branches]
                        if branch.id not in user_branch_ids:
                            errors.append(f"Row {row_num}: No access to branch '{row[1]}'")
                            continue

                    cat_name = str(row[7] or "").strip().lower()
                    cat = cats_map.get(cat_name)
                    if not cat:
                        errors.append(f"Row {row_num}: Unknown category '{row[7]}'")
                        continue

                    qty = int(row[8] or 1)
                    rate = Decimal(str(row[9] or 0))
                    amount = qty * rate

                    item = PaymentRequestItem(
                        description=str(row[6] or "").strip(),
                        category_id=cat.id,
                        quantity=qty,
                        rate=rate,
                        amount=amount,
                    )
                    pr = PaymentRequest(
                        reference=PaymentRequest.generate_reference(),
                        date=row_date,
                        branch_id=branch.id,
                        beneficiary_name=str(row[2] or "").strip(),
                        beneficiary_account=str(row[3] or "").strip(),
                        beneficiary_bank=str(row[4] or "").strip(),
                        bank_code=str(row[5] or "").strip() or None,
                        requested_amount=amount,
                        submitted_by_id=current_user.id,
                    )
                    pr.items = [item]
                    db.session.add(pr)
                    db.session.flush()
                    created += 1

                except Exception as row_err:
                    errors.append(f"Row {row_num}: {str(row_err)}")
                    continue

            db.session.commit()
            if created:
                flash(f"Successfully created {created} payment request(s).", "success")
            for err in errors[:8]:
                flash(err, "warning")
            return redirect(url_for("requests.list_requests"))

        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            flash(f"Error processing file: {str(e)}", "danger")

    return render_template("bulk_upload.html")


# ── Template Download ─────────────────────────────────────────────────────────

@requests_bp.route("/requests/template")
@login_required
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Requests"

    header_fill = PatternFill("solid", fgColor="3A9FD8")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Date (YYYY-MM-DD)", "Branch", "Beneficiary Name", "Account Number",
        "Bank", "Bank Code (opt)", "Description", "Category", "Quantity", "Rate (₦)", "Amount (₦)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    branches = Branch.query.filter_by(is_active=True).all()
    branch_name = branches[0].name if branches else "Hopestone Hospital (Ikeja)"
    sample = [str(date.today()), branch_name, "Sample Vendor", "1234567890",
              "Sterling Bank", "", "Paracetamol 500mg x100", "Drugs", 2, 3500, 7000]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "VALID BRANCHES:"
    ws2["A1"].font = Font(bold=True, color="1A3A5C")
    for i, b in enumerate(branches, 2):
        ws2.cell(row=i, column=1, value=b.name)

    categories = Category.query.filter_by(is_active=True).order_by(Category.cost_type, Category.name).all()
    ws2["C1"] = "VALID CATEGORIES:"
    ws2["C1"].font = Font(bold=True, color="1A3A5C")
    for i, c in enumerate(categories, 2):
        ws2.cell(row=i, column=3, value=c.name)
        ws2.cell(row=i, column=4, value=f"({c.cost_type.replace('_', ' ').title()})")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="hopestone_payment_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Export ────────────────────────────────────────────────────────────────────

@requests_bp.route("/requests/export")
@login_required
def export_requests():
    q = _eager_pr()
    if not current_user.is_mds:
        q = q.filter_by(submitted_by_id=current_user.id)

    status_filter = request.args.get("status", "approved")
    branch_filter = request.args.get("branch_id", "")
    month_filter = request.args.get("month", "")

    if status_filter:
        q = q.filter_by(status=status_filter)
    if branch_filter:
        q = q.filter_by(branch_id=int(branch_filter))
    if month_filter:
        try:
            year, month = month_filter.split("-")
            q = q.filter(
                extract("year", PaymentRequest.created_at) == int(year),
                extract("month", PaymentRequest.created_at) == int(month),
            )
        except Exception:
            pass

    records = q.order_by(PaymentRequest.date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hopestone Payments"

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Reference", "Date", "Branch", "Beneficiary Name", "Account Number",
        "Bank", "Bank Code", "Requested (₦)", "Approved (₦)", "Status",
        "Category", "Parent Item", "Description", "Note", "Qty", "Rate (₦)", "Amount (₦)",
        "Upload Status", "Submitted By",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for pr in records:
        # Collect leaf items only (no double-counting of parent roll-ups)
        for item in pr.items:
            # Skip parent roll-up rows (items that have children)
            if item.children:
                continue
            ws.cell(row=row, column=1, value=pr.reference)
            ws.cell(row=row, column=2, value=str(pr.date))
            ws.cell(row=row, column=3, value=pr.branch.name if pr.branch else "")
            ws.cell(row=row, column=4, value=pr.beneficiary_name)
            ws.cell(row=row, column=5, value=pr.beneficiary_account)
            ws.cell(row=row, column=6, value=pr.beneficiary_bank)
            ws.cell(row=row, column=7, value=pr.bank_code or "")
            ws.cell(row=row, column=8, value=float(pr.requested_amount))
            ws.cell(row=row, column=9, value=float(pr.approved_amount) if pr.approved_amount else "")
            ws.cell(row=row, column=10, value=pr.status.title())
            ws.cell(row=row, column=11, value=item.category.name if item.category else "")
            ws.cell(row=row, column=12, value=item.parent_item.description if item.parent_id else "")
            ws.cell(row=row, column=13, value=item.description)
            ws.cell(row=row, column=14, value=item.note or "")
            ws.cell(row=row, column=15, value=item.quantity)
            ws.cell(row=row, column=16, value=float(item.rate))
            ws.cell(row=row, column=17, value=float(item.amount))
            ws.cell(row=row, column=18, value=pr.upload_status.replace("_", " ").title())
            ws.cell(row=row, column=19, value=pr.submitter.name if pr.submitter else "")
            row += 1

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"hopestone_payments_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
