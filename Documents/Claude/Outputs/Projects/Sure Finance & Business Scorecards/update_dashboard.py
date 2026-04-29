#!/usr/bin/env python3
"""Sure Finance Weekly Dashboard Updater v2.1
Changes vs v1:
  - Tests = actual investigation count (each comma-separated item in Investigations column)
  - Cases = number of case rows (separate metric)
  - Unique Referrers (nunique) everywhere
  - Agent Performance section (new)
  - Revenue vs Target comparison (Q1-Q4 targets embedded)
  - Centre name normalization
"""
import sys, os, re, json, datetime, shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE = os.path.dirname(os.path.abspath(__file__))

# ── TARGETS (2026) ──────────────────────────────────────────────────────────
TARGETS = {
    'Ilasa Main Centre, Lagos': {
        'short': 'Ilasa', 'color': '2E86AB',
        'monthly': [5_000_000, 6_100_000, 7_000_000, 7_950_000],
        'daily':   [  200_000,   244_000,   280_000,   318_000],
    },
    'OAUTH Ilesa Centre': {
        'short': 'OAUTH', 'color': '8E44AD',
        'monthly': [5_500_000, 7_000_000, 7_500_000, 8_000_000],
        'daily':   [  220_000,   280_000,   300_000,   320_000],
    },
    'Palm Avenue, Lagos': {
        'short': 'Palm Ave', 'color': '16A085',
        'monthly': [2_500_000, 3_500_000, 4_000_000, 4_410_000],
        'daily':   [  100_000,   140_000,   160_000,   176_400],
    },
    'Sure Ilesha, Osun': {
        'short': 'Sure Ilesha', 'color': 'E67E22',
        'monthly': [4_250_000, 6_000_000, 7_500_000, 8_500_000],
        'daily':   [  170_000,   240_000,   300_000,   340_000],
    },
    'Ikeja Lagos Centre': {
        'short': 'Ikeja', 'color': '2196F3',
        'monthly': [2_000_000, 3_000_000, 5_250_000, 7_250_000],
        'daily':   [   80_000,   120_000,   210_000,   290_000],
    },
}

WORKSPACE = os.path.dirname(os.path.abspath(__file__))

def fill(h): return PatternFill('solid', start_color=h, end_color=h)
def fnt(bold=False, sz=10, color='000000', italic=False):
    return Font(name='Arial', bold=bold, size=sz, color=color, italic=italic)
def aln(h='left', v='center'):
    return Alignment(horizontal=h, vertical=v)
def bdr():
    s = Side(style='thin', color='BFBFBF')
    return Border(top=s, bottom=s, left=s, right=s)

def normalize_centre(name):
    if pd.isna(name): return 'Unknown'
    return str(name).strip()

def normalize_agent(name):
    if pd.isna(name) or str(name).strip() == '':
        return 'Walk-in / Unassigned'
    return re.sub(r'\s+', ' ', str(name).strip())

def count_investigations(val):
    """Count individual investigations (comma-separated items)."""
    if pd.isna(val) or str(val).strip() == '':
        return 0
    items = [i.strip() for i in str(val).split(',') if i.strip()]
    return max(len(items), 1)  # at least 1 if field has any content

def get_quarter(date):
    return (date.month - 1) // 3

def ctarget(centre, qi, key):
    cfg = TARGETS.get(centre, {})
    vals = cfg.get(key, [])
    if not vals or qi >= len(vals): return None
    return vals[qi]

# ── PROCESS CSV ───────────────────────────────────────────────────────────────
def process_csv(csv_path):
    print(f"Reading: {csv_path}")
    df = pd.read_csv(csv_path)
    df['Date'] = pd.to_datetime(df['Date'])
    df_active = df[df['Canceled'] == False].copy()
    df_active['Collection centre'] = df_active['Collection centre'].apply(normalize_centre)
    df_active['Agent'] = df_active['Agent'].apply(normalize_agent)

    # Count individual investigations per row
    df_active['inv_count'] = df_active['Investigations'].apply(count_investigations)

    dates = sorted(df_active['Date'].unique())
    if not dates:
        raise ValueError("No active cases found.")
    date_strs   = [d.strftime('%Y-%m-%d') for d in dates]
    date_labels = [d.strftime('%a %d %b') for d in dates]
    num_days = len(dates)
    min_d = min(dates); max_d = max(dates)
    qi = get_quarter(min_d)
    week_label = f"{min_d.strftime('%a %d')} \u2013 {max_d.strftime('%a %d %b %Y')}"

    centres = sorted(df_active['Collection centre'].unique().tolist())

    # Per-centre targets
    ctargs = {}
    for c in centres:
        d = ctarget(c, qi, 'daily')
        m = ctarget(c, qi, 'monthly')
        cfg = TARGETS.get(c, {})
        ctargs[c] = {
            'short':   cfg.get('short', c.split(',')[0]),
            'color':   '#' + cfg.get('color', '888888'),
            'daily':   d,
            'weekly':  d * num_days if d else None,
            'monthly': m,
        }

    # Daily aggregation
    daily = df_active.groupby(['Date', 'Collection centre']).agg(
        investigations=('inv_count', 'sum'),   # actual test count
        cases=('Case Id', 'count'),            # case row count
        patients=('Reg. no.', 'nunique'),
        revenue=('Total Fee', 'sum'),
        paid=('Fee Paid', 'sum'),
        due=('Fee Due', 'sum'),
        unique_referrers=('Referrer', 'nunique'),
    ).reset_index()
    daily['Date'] = daily['Date'].astype(str)

    # Case type mix
    ct = df_active.groupby(['Date', 'Collection centre', 'Case Type']).agg(
        count=('Case Id', 'count'),
        investigations=('inv_count', 'sum'),
    ).reset_index()
    ct['Date'] = ct['Date'].astype(str)

    # Weekly aggregation
    weekly = df_active.groupby('Collection centre').agg(
        investigations=('inv_count', 'sum'),
        cases=('Case Id', 'count'),
        patients=('Reg. no.', 'nunique'),
        revenue=('Total Fee', 'sum'),
        paid=('Fee Paid', 'sum'),
        due=('Fee Due', 'sum'),
        unique_referrers=('Referrer', 'nunique'),
    ).reset_index()

    # Top referrers
    top_ref = df_active.groupby(['Collection centre', 'Referrer']).agg(
        cases=('Case Id', 'count'),
        investigations=('inv_count', 'sum'),
        revenue=('Total Fee', 'sum'),
    ).reset_index().sort_values(['Collection centre', 'cases'], ascending=[True, False])

    # Agent performance
    agent_total = df_active.groupby('Agent').agg(
        investigations=('inv_count', 'sum'),
        cases=('Case Id', 'count'),
        patients=('Reg. no.', 'nunique'),
        revenue=('Total Fee', 'sum'),
        paid=('Fee Paid', 'sum'),
    ).reset_index().sort_values('investigations', ascending=False)

    agent_centre = df_active.groupby(['Agent', 'Collection centre']).agg(
        investigations=('inv_count', 'sum'),
        cases=('Case Id', 'count'),
        revenue=('Total Fee', 'sum'),
        paid=('Fee Paid', 'sum'),
    ).reset_index()

    total_rev  = int(df_active['Total Fee'].sum())
    total_paid = int(df_active['Fee Paid'].sum())
    total_inv  = int(df_active['inv_count'].sum())
    total_cases = int(df_active.shape[0])

    data = {
        'week_label':   week_label,
        'generated':    datetime.date.today().strftime('%d %b %Y'),
        'quarter':      f'Q{qi+1} 2026',
        'num_days':     num_days,
        'centres':      centres,
        'dates':        date_strs,
        'date_labels':  date_labels,
        'daily':        daily.to_dict(orient='records'),
        'case_types':   ct.to_dict(orient='records'),
        'weekly':       weekly.to_dict(orient='records'),
        'top_referrers': top_ref.head(120).to_dict(orient='records'),
        'agent_total':   agent_total.to_dict(orient='records'),
        'agent_centre':  agent_centre.to_dict(orient='records'),
        'centre_targets': ctargs,
        'summary': {
            'total_investigations':    total_inv,
            'total_cases':             total_cases,
            'total_patients':          int(df_active['Reg. no.'].nunique()),
            'total_unique_referrers':  int(df_active['Referrer'].nunique()),
            'total_revenue':           total_rev,
            'total_paid':              total_paid,
            'total_due':               int(df_active['Fee Due'].sum()),
            'total_discount':          int(df_active['Discount'].sum()),
            'cancelled':               int((df['Canceled'] == True).sum()),
            'collection_rate':         round(total_paid / total_rev * 100, 1) if total_rev > 0 else 0,
        }
    }
    return data, df_active, df, week_label


# ── EXCEL ─────────────────────────────────────────────────────────────────────
NAVY='1F3864'; TEAL='2E86AB'; WHITE='FFFFFF'; LGRAY='F0F4F8'
CENTRE_COLORS = ['2E86AB','8E44AD','16A085','E67E22','E74C3C','F39C12']

def build_excel(data, df_active, week_label):
    wb = Workbook()
    wb.remove(wb.active)

    S       = data['summary']
    centres = data['centres']
    weekly  = {r['Collection centre']: r for r in data['weekly']}
    ctargs  = data['centre_targets']

    # ── Sheet 1: Dashboard ──────────────────────────────────────────────────
    ws = wb.create_sheet('📊 Dashboard')
    ws.column_dimensions['A'].width = 32
    for col in 'BCDEFGHIJK':
        ws.column_dimensions[col].width = 18

    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = 'SURE FINANCE & BUSINESS — WEEKLY PERFORMANCE DASHBOARD'
    c.font = fnt(bold=True, sz=14, color=WHITE)
    c.fill = fill(NAVY); c.alignment = aln('center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:K2')
    c = ws['A2']
    c.value = f'Week: {week_label}  |  Generated: {data["generated"]}  |  {data["quarter"]} targets applied'
    c.font = fnt(sz=10, color=WHITE, italic=True)
    c.fill = fill(TEAL); c.alignment = aln('center')

    # KPI block
    ws['A4'] = 'KEY METRICS'; ws['A4'].font = fnt(bold=True, sz=11, color=NAVY)
    kpis = [
        ('Total Investigations (tests)', S['total_investigations']),
        ('Total Case Rows',              S['total_cases']),
        ('Unique Patients',              S['total_patients']),
        ('Unique Referrers',             S['total_unique_referrers']),
        ('Total Revenue (₦)',            S['total_revenue']),
        ('Fee Collected (₦)',            S['total_paid']),
        ('Fee Due (₦)',                  S['total_due']),
        ('Discount Given (₦)',           S['total_discount']),
        ('Collection Rate',              f"{S['collection_rate']}%"),
        ('Cancelled Cases',              S['cancelled']),
    ]
    for i, (label, val) in enumerate(kpis):
        row = 5 + i
        ws.cell(row, 1, label).font = fnt(bold=True, sz=10)
        c2 = ws.cell(row, 2, val)
        c2.font = fnt(bold=True, sz=10, color=NAVY)
        if isinstance(val, int) and val > 999:
            c2.number_format = '₦#,##0'

    # Centre summary header
    hrow = 17
    headers = ['Collection Centre', 'Investigations', 'Cases', 'Patients', 'Unique Referrers',
               'Total Revenue (₦)', 'Paid (₦)', 'Due (₦)',
               'Weekly Target (₦)', 'Achievement %', 'Collection %']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(hrow, ci, h)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(NAVY); c.alignment = aln('center'); c.border = bdr()

    for ri, centre in enumerate(centres):
        row = hrow + 1 + ri
        r   = weekly.get(centre, {})
        tgt = ctargs.get(centre, {})
        rev = r.get('revenue', 0)
        wt  = tgt.get('weekly')
        achv = f"{rev/wt*100:.1f}%" if wt and rev else 'No target'
        cr   = f"{r.get('paid',0)/rev*100:.1f}%" if rev else '–'
        col_hex = CENTRE_COLORS[ri % len(CENTRE_COLORS)]

        vals = [centre, r.get('investigations',''), r.get('cases',''), r.get('patients',''),
                r.get('unique_referrers',''), rev, r.get('paid',''),
                r.get('due',''), wt or '', achv, cr]
        for ci2, v in enumerate(vals, 1):
            c = ws.cell(row, ci2, v)
            c.font = fnt(sz=10, bold=(ci2==1), color=(col_hex if ci2==1 else '000000'))
            c.border = bdr(); c.alignment = aln('center' if ci2 > 1 else 'left')
            if ci2 in (6,7,8,9) and isinstance(v, (int,float)):
                c.number_format = '₦#,##0'

    # Totals row
    trow = hrow + 1 + len(centres)
    total_wt = sum(ctargs[c].get('weekly') or 0 for c in centres if ctargs[c].get('weekly'))
    tot_rev  = S['total_revenue']
    tot_achv = f"{tot_rev/total_wt*100:.1f}%" if total_wt else '–'
    totals = ['TOTAL', S['total_investigations'], S['total_cases'], S['total_patients'],
              S['total_unique_referrers'], tot_rev, S['total_paid'],
              S['total_due'], total_wt or '', tot_achv, f"{S['collection_rate']}%"]
    for ci2, v in enumerate(totals, 1):
        c = ws.cell(trow, ci2, v)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(TEAL); c.border = bdr()
        c.alignment = aln('center' if ci2 > 1 else 'left')
        if ci2 in (6,7,8,9) and isinstance(v, (int,float)):
            c.number_format = '₦#,##0'

    ws.freeze_panes = 'B5'

    # ── Sheet 2: Daily Performance ──────────────────────────────────────────
    ws2 = wb.create_sheet('📅 Daily Performance')
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 28
    for col in 'CDEFGHIJKL':
        ws2.column_dimensions[col].width = 16

    ws2.merge_cells('A1:L1')
    c = ws2['A1']
    c.value = f'Daily Performance — {week_label}'
    c.font = fnt(bold=True, sz=12, color=WHITE); c.fill = fill(NAVY)
    c.alignment = aln('center'); ws2.row_dimensions[1].height = 22

    h2 = ['Date', 'Collection Centre', 'Investigations', 'Cases', 'Patients',
          'Unique Referrers', 'Total Fee (₦)', 'Paid (₦)', 'Due (₦)',
          'Daily Target (₦)', 'Achievement %', 'Collection %']
    for ci, h in enumerate(h2, 1):
        c = ws2.cell(2, ci, h)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(TEAL); c.alignment = aln('center'); c.border = bdr()

    daily_rows = sorted(data['daily'], key=lambda r: (r['Date'], r['Collection centre']))
    for ri, r in enumerate(daily_rows):
        row = 3 + ri
        centre  = r['Collection centre']
        tgt_ci  = centres.index(centre) if centre in centres else 0
        col_hex = CENTRE_COLORS[tgt_ci % len(CENTRE_COLORS)]
        dt  = ctargs.get(centre, {}).get('daily')
        rev = r.get('revenue', 0)
        achv = f"{rev/dt*100:.1f}%" if dt and rev else ('No target' if not dt else '–')
        cr   = f"{r.get('paid',0)/rev*100:.1f}%" if rev else '–'
        vals2 = [r['Date'], centre, r.get('investigations',''), r.get('cases',''),
                 r.get('patients',''), r.get('unique_referrers',''),
                 rev, r.get('paid',''), r.get('due',''), dt or '', achv, cr]
        for ci2, v in enumerate(vals2, 1):
            c = ws2.cell(row, ci2, v)
            c.font = fnt(sz=10, color=(col_hex if ci2==2 else '000000'), bold=(ci2==2))
            c.border = bdr(); c.alignment = aln('center' if ci2 > 2 else 'left')
            if ci2 in (7,8,9,10) and isinstance(v, (int,float)):
                c.number_format = '₦#,##0'
    ws2.freeze_panes = 'C3'

    # ── Sheet 3: Weekly Summary ─────────────────────────────────────────────
    ws3 = wb.create_sheet('🏥 Weekly Summary')
    ws3.column_dimensions['A'].width = 28
    for col in 'BCDEFGHIJKL':
        ws3.column_dimensions[col].width = 16

    ws3.merge_cells('A1:L1')
    c = ws3['A1']
    c.value = f'Weekly Centre Summary vs {data["quarter"]} Target — {week_label}'
    c.font = fnt(bold=True, sz=12, color=WHITE); c.fill = fill(NAVY)
    c.alignment = aln('center'); ws3.row_dimensions[1].height = 22

    h3 = ['Collection Centre', 'Investigations', 'Cases', 'Patients', 'Unique Referrers',
          'Total Revenue (₦)', 'Paid (₦)', 'Due (₦)',
          'Daily Target (₦)', 'Weekly Target (₦)', 'Achievement %', 'Collection %']
    for ci, h in enumerate(h3, 1):
        c = ws3.cell(2, ci, h)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(TEAL); c.alignment = aln('center'); c.border = bdr()

    for ri, centre in enumerate(centres):
        row = 3 + ri
        r   = weekly.get(centre, {})
        tgt = ctargs.get(centre, {})
        rev = r.get('revenue', 0)
        dt  = tgt.get('daily'); wt = tgt.get('weekly')
        achv = f"{rev/wt*100:.1f}%" if wt and rev else 'No target'
        cr   = f"{r.get('paid',0)/rev*100:.1f}%" if rev else '–'
        col_hex = CENTRE_COLORS[ri % len(CENTRE_COLORS)]
        vals3 = [centre, r.get('investigations',''), r.get('cases',''), r.get('patients',''),
                 r.get('unique_referrers',''), rev, r.get('paid',''),
                 r.get('due',''), dt or '', wt or '', achv, cr]
        for ci2, v in enumerate(vals3, 1):
            c = ws3.cell(row, ci2, v)
            c.font = fnt(sz=10, color=(col_hex if ci2==1 else '000000'), bold=(ci2==1))
            c.border = bdr(); c.alignment = aln('center' if ci2 > 1 else 'left')
            if ci2 in (6,7,8,9,10) and isinstance(v, (int,float)):
                c.number_format = '₦#,##0'
    ws3.freeze_panes = 'B3'

    # ── Sheet 4: Referral Analysis ──────────────────────────────────────────
    ws4 = wb.create_sheet('🔗 Referral Analysis')
    ws4.column_dimensions['A'].width = 26
    ws4.column_dimensions['B'].width = 34
    ws4.column_dimensions['C'].width = 14
    ws4.column_dimensions['D'].width = 16
    ws4.column_dimensions['E'].width = 18

    ws4.merge_cells('A1:E1')
    c = ws4['A1']
    c.value = f'Referral Analysis — {week_label}'
    c.font = fnt(bold=True, sz=12, color=WHITE); c.fill = fill(NAVY)
    c.alignment = aln('center'); ws4.row_dimensions[1].height = 22

    for ci, h in enumerate(['Collection Centre', 'Referrer / Doctor', 'Cases', 'Investigations', 'Revenue (₦)'], 1):
        c = ws4.cell(2, ci, h)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(TEAL); c.alignment = aln('center'); c.border = bdr()

    for ri, r in enumerate(data['top_referrers']):
        row = 3 + ri
        centre  = r['Collection centre']
        tgt_ci  = centres.index(centre) if centre in centres else 0
        col_hex = CENTRE_COLORS[tgt_ci % len(CENTRE_COLORS)]
        for ci2, v in enumerate([centre, r['Referrer'], r['cases'], r.get('investigations',''), r['revenue']], 1):
            c = ws4.cell(row, ci2, v)
            c.font = fnt(sz=10, color=(col_hex if ci2==1 else '000000'), bold=(ci2==1))
            c.border = bdr(); c.alignment = aln('left' if ci2<=2 else 'center')
            if ci2 == 5: c.number_format = '₦#,##0'
    ws4.freeze_panes = 'A3'

    # ── Sheet 5: Agent Performance ──────────────────────────────────────────
    ws5 = wb.create_sheet('👥 Agent Performance')
    for col, w in zip('ABCDEFG', [26,16,12,14,18,18,14]):
        ws5.column_dimensions[col].width = w

    ws5.merge_cells('A1:G1')
    c = ws5['A1']
    c.value = f'Agent Performance — {week_label}'
    c.font = fnt(bold=True, sz=12, color=WHITE); c.fill = fill(NAVY)
    c.alignment = aln('center'); ws5.row_dimensions[1].height = 22

    for ci, h in enumerate(['Agent', 'Investigations', 'Cases', 'Patients',
                             'Total Revenue (₦)', 'Collected (₦)', 'Collection %'], 1):
        c = ws5.cell(2, ci, h)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(TEAL); c.alignment = aln('center'); c.border = bdr()

    for ri, r in enumerate(data['agent_total']):
        row = 3 + ri
        rev  = r.get('revenue', 0)
        paid = r.get('paid', 0)
        cr   = f"{paid/rev*100:.1f}%" if rev else '–'
        for ci2, v in enumerate([r['Agent'], r.get('investigations',''), r.get('cases',''),
                                  r.get('patients',''), rev, paid, cr], 1):
            c = ws5.cell(row, ci2, v)
            c.font = fnt(sz=10, bold=(ci2==1))
            c.border = bdr(); c.alignment = aln('left' if ci2==1 else 'center')
            if ci2 in (5,6) and isinstance(v,(int,float)):
                c.number_format = '₦#,##0'

    # Agent × Centre breakdown sub-table
    br = 3 + len(data['agent_total']) + 2
    ws5.cell(br, 1, 'Agent × Centre Breakdown').font = fnt(bold=True, sz=11, color=NAVY)
    br += 1
    for ci, h in enumerate(['Agent', 'Collection Centre', 'Investigations', 'Cases',
                             'Revenue (₦)', 'Collected (₦)'], 1):
        c = ws5.cell(br, ci, h)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(NAVY); c.alignment = aln('center'); c.border = bdr()

    ac_sorted = sorted(data['agent_centre'], key=lambda r: (-r.get('investigations',0), r['Agent']))
    for ri, r in enumerate(ac_sorted):
        row = br + 1 + ri
        centre  = r['Collection centre']
        tgt_ci  = centres.index(centre) if centre in centres else 0
        col_hex = CENTRE_COLORS[tgt_ci % len(CENTRE_COLORS)]
        for ci2, v in enumerate([r['Agent'], centre, r.get('investigations',''), r.get('cases',''),
                                  r.get('revenue',0), r.get('paid',0)], 1):
            c = ws5.cell(row, ci2, v)
            c.font = fnt(sz=10, color=(col_hex if ci2==2 else '000000'))
            c.border = bdr(); c.alignment = aln('left' if ci2<=2 else 'center')
            if ci2 in (5,6) and isinstance(v,(int,float)):
                c.number_format = '₦#,##0'
    ws5.freeze_panes = 'A3'

    # ── Sheet 6: Raw Data ───────────────────────────────────────────────────
    ws6 = wb.create_sheet('📋 Raw Data')
    raw_cols = list(df_active.columns)
    for ci, h in enumerate(raw_cols, 1):
        c = ws6.cell(1, ci, h)
        c.font = fnt(bold=True, sz=10, color=WHITE)
        c.fill = fill(NAVY); c.alignment = aln('center')
    for ri, row_data in enumerate(df_active.itertuples(index=False), 2):
        for ci, v in enumerate(row_data, 1):
            ws6.cell(ri, ci, str(v) if not isinstance(v,(int,float,type(None))) else v)
    ws6.freeze_panes = 'A2'

    safe_label = week_label.replace(' \u2013 ',' to ').replace(' ','_')[:30]
    out_path = os.path.join(WORKSPACE, f"Sure_Weekly_Dashboard_{safe_label}.xlsx")
    wb.save(out_path)
    print(f"Excel saved: {out_path}")
    return out_path


# ── HTML ──────────────────────────────────────────────────────────────────────
def build_html(data):
    data_json = json.dumps(data, indent=2, default=str)

    html_parts = []
    html_parts.append("""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Sure Finance \u2013 Weekly Performance Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root{--navy:#1F3864;--teal:#2E86AB;--green:#27AE60;--amber:#E67E22;--red:#E74C3C;
      --purple:#8E44AD;--dgreen:#16A085;--gold:#F39C12;--bg:#F0F4F8;--card:#fff;--border:#DDE3EA;}
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:Arial,sans-serif;background:var(--bg);color:#2C3E50;font-size:13px;}
.header{background:var(--navy);color:#fff;padding:16px 24px;display:flex;justify-content:space-between;align-items:center;}
.header h1{font-size:17px;font-weight:700;letter-spacing:.4px;}
.header .meta{font-size:11px;opacity:.75;text-align:right;line-height:1.6;}
.week-badge{background:var(--teal);color:#fff;padding:4px 12px;border-radius:20px;font-size:11px;font-weight:700;margin-top:4px;display:inline-block;}
.filter-bar{background:#fff;border-bottom:2px solid var(--border);padding:0 24px;display:flex;gap:0;overflow-x:auto;}
.tab{padding:10px 18px;cursor:pointer;font-size:12px;font-weight:600;color:#666;border-bottom:3px solid transparent;transition:all .2s;white-space:nowrap;}
.tab:hover{color:var(--navy);background:#F8FAFC;}.tab.active{color:var(--navy);border-bottom-color:var(--teal);}
.main{padding:20px 24px;max-width:1500px;margin:0 auto;}
.section-title{font-size:13px;font-weight:700;color:var(--navy);text-transform:uppercase;
  letter-spacing:.8px;margin:20px 0 10px;padding-bottom:6px;border-bottom:2px solid var(--border);}
.kpi-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:12px;margin-bottom:20px;}
.kpi-card{background:var(--card);border-radius:10px;padding:14px 16px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);border-top:4px solid var(--teal);text-align:center;}
.kpi-label{font-size:10px;text-transform:uppercase;letter-spacing:.6px;color:#888;font-weight:700;}
.kpi-value{font-size:22px;font-weight:800;color:var(--navy);margin:6px 0 2px;}
.kpi-sub{font-size:10px;color:#aaa;}
.charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;}
.chart-card{background:var(--card);border-radius:10px;padding:16px;box-shadow:0 1px 4px rgba(0,0,0,.08);}
.chart-title{font-size:12px;font-weight:700;color:var(--navy);margin-bottom:12px;}
.chart-wrap{position:relative;height:220px;}
.table-card{background:var(--card);border-radius:10px;padding:16px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:16px;overflow-x:auto;}
table{width:100%;border-collapse:collapse;}
thead tr{background:var(--navy);color:#fff;}
thead th{padding:8px 10px;font-size:11px;font-weight:700;text-align:center;white-space:nowrap;}
thead th:first-child{text-align:left;}
tbody tr:nth-child(even){background:#F7FAFC;}
tbody tr:hover{background:#EBF5FB;}
tbody td{padding:7px 10px;border-bottom:1px solid var(--border);font-size:12px;text-align:center;white-space:nowrap;}
tbody td:first-child{text-align:left;font-weight:600;}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;}
.badge-green{background:#D5F5E3;color:#1E8449;}.badge-amber{background:#FCF3CF;color:#9A7D0A;}
.badge-red{background:#FADBD8;color:#922B21;}.badge-gray{background:#EAEDED;color:#555;}
.dot{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:5px;vertical-align:middle;}
.note{background:#FEF9E7;border-left:4px solid var(--gold);padding:8px 14px;
  font-size:11px;color:#7D6608;border-radius:0 6px 6px 0;margin-bottom:16px;}
.footer{text-align:center;padding:16px;font-size:10px;color:#aaa;background:#fff;
  border-top:1px solid var(--border);margin-top:24px;}
@media(max-width:900px){.kpi-grid{grid-template-columns:repeat(3,1fr);}.charts-grid{grid-template-columns:1fr;}}
</style></head><body>
<div class="header">
  <div><h1>&#128202; SURE FINANCE &amp; BUSINESS \u2014 WEEKLY PERFORMANCE DASHBOARD</h1>
  <div style="font-size:11px;opacity:.7;margin-top:3px" id="centreList"></div></div>
  <div class="meta"><div><span class="week-badge" id="weekBadge"></span></div>
  <div style="margin-top:4px">Generated: <span id="genDate"></span></div></div>
</div>
<div class="filter-bar" id="filterBar">
  <div class="tab active" onclick="filterCentre('ALL',this)">&#127970; All Centres</div>
</div>
<div class="main">
<div class="note" id="topNote"></div>
<div class="section-title">Key Performance Indicators \u2014 Week to Date</div>
<div class="kpi-grid">
  <div class="kpi-card" style="border-top-color:var(--teal)"><div class="kpi-label">Total Tests</div><div class="kpi-value" id="kpi-inv">\u2013</div><div class="kpi-sub">Individual investigations</div></div>
  <div class="kpi-card" style="border-top-color:var(--dgreen)"><div class="kpi-label">Patients</div><div class="kpi-value" id="kpi-patients">\u2013</div><div class="kpi-sub">Unique Reg. nos</div></div>
  <div class="kpi-card" style="border-top-color:var(--purple)"><div class="kpi-label">Unique Referrers</div><div class="kpi-value" id="kpi-referrers">\u2013</div><div class="kpi-sub">Distinct doctors/sources</div></div>
  <div class="kpi-card" style="border-top-color:var(--navy)"><div class="kpi-label">Total Revenue</div><div class="kpi-value" id="kpi-revenue">\u20a6\u2013</div><div class="kpi-sub" id="kpi-rev-sub">Total fee billed</div></div>
  <div class="kpi-card" style="border-top-color:var(--green)"><div class="kpi-label">Fee Collected</div><div class="kpi-value" id="kpi-paid">\u20a6\u2013</div><div class="kpi-sub">Amount paid</div></div>
  <div class="kpi-card" style="border-top-color:var(--amber)"><div class="kpi-label">Fee Due</div><div class="kpi-value" id="kpi-due">\u20a6\u2013</div><div class="kpi-sub">Outstanding</div></div>
  <div class="kpi-card" style="border-top-color:var(--red)"><div class="kpi-label">Collection Rate</div><div class="kpi-value" id="kpi-rate">\u2013</div><div class="kpi-sub">Paid / Total fee</div></div>
</div>
<div class="section-title">Performance Charts</div>
<div class="charts-grid">
  <div class="chart-card"><div class="chart-title">&#128197; Daily Investigations by Centre</div><div class="chart-wrap"><canvas id="chartDailyInv"></canvas></div></div>
  <div class="chart-card"><div class="chart-title">&#128176; Daily Revenue by Centre (\u20a6)</div><div class="chart-wrap"><canvas id="chartDailyRev"></canvas></div></div>
  <div class="chart-card"><div class="chart-title">&#127919; Weekly Revenue vs Target (\u20a6)</div><div class="chart-wrap"><canvas id="chartVsTarget"></canvas></div></div>
  <div class="chart-card"><div class="chart-title">&#128302; Investigation Type Mix by Centre</div><div class="chart-wrap"><canvas id="chartCaseTypes"></canvas></div></div>
</div>
<div style="margin-top:16px" class="charts-grid">
  <div class="chart-card"><div class="chart-title">&#128179; Weekly Revenue vs Collected (\u20a6)</div><div class="chart-wrap"><canvas id="chartWeeklyRev"></canvas></div></div>
  <div class="chart-card"><div class="chart-title">&#128200; Daily Collection Rate (%)</div><div class="chart-wrap"><canvas id="chartCollRate"></canvas></div></div>
</div>
<div class="section-title">Daily Performance Summary</div>
<div class="table-card"><table>
<thead><tr><th>Date</th><th>Centre</th><th>Investigations</th><th>Cases</th><th>Patients</th>
<th>Unique Referrers</th><th>Total Fee (\u20a6)</th><th>Paid (\u20a6)</th><th>Due (\u20a6)</th>
<th>Daily Target (\u20a6)</th><th>Achievement</th><th>Collection %</th></tr></thead>
<tbody id="dailyTbody"></tbody></table></div>
<div class="section-title">Weekly Centre Summary vs Target</div>
<div class="table-card"><table>
<thead><tr><th>Collection Centre</th><th>Investigations</th><th>Cases</th><th>Patients</th>
<th>Unique Referrers</th><th>Total Fee (\u20a6)</th><th>Paid (\u20a6)</th><th>Due (\u20a6)</th>
<th>Weekly Target (\u20a6)</th><th>Achievement %</th><th>Avg Fee/Case</th><th>Collection %</th>
</tr></thead><tbody id="weeklyTbody"></tbody></table></div>
<div class="section-title">Top Referrers by Collection Centre</div>
<div class="table-card"><table>
<thead><tr><th>#</th><th>Centre</th><th>Referrer / Doctor</th>
<th>Cases</th><th>Revenue (\u20a6)</th><th>% of Centre</th></tr></thead>
<tbody id="refTbody"></tbody></table></div>
<div class="section-title">Agent Performance</div>
<div class="table-card"><table>
<thead><tr><th>#</th><th>Agent</th><th>Investigations</th><th>Cases</th><th>Patients</th>
<th>Revenue (\u20a6)</th><th>Collected (\u20a6)</th><th>Collection %</th></tr></thead>
<tbody id="agentTbody"></tbody></table></div>
</div>
<div class="footer">Sure Finance &amp; Business Scorecards &nbsp;|&nbsp; Auto-generated from LabSmartLIS export</div>
""")

    html_parts.append('<script>\nconst DATA = ')
    html_parts.append(data_json)
    html_parts.append(""";
const COLOURS=['#2E86AB','#8E44AD','#16A085','#E67E22','#E74C3C','#F39C12'];
let activeFilter='ALL';
const charts={};
const fmt=n=>(n==null)?'\u2013':'\u20a6'+Math.round(n).toLocaleString();
const fmtN=n=>(n==null)?'\u2013':Number(n).toLocaleString();
function badge(val,hi,lo){
  if(val==null||val==='\u2013')return'<span class="badge badge-gray">\u2013</span>';
  const n=parseFloat(val);
  if(isNaN(n))return`<span class="badge badge-gray">${val}</span>`;
  const cls=n>=hi?'badge-green':n>=lo?'badge-amber':'badge-red';
  return`<span class="badge ${cls}">${val}</span>`;
}
function achvBadge(actual,target){
  if(!target)return'<span class="badge badge-gray">No target</span>';
  if(!actual)return'<span class="badge badge-gray">\u2013</span>';
  const p=actual/target*100;
  const cls=p>=100?'badge-green':p>=80?'badge-amber':'badge-red';
  return`<span class="badge ${cls}">${p.toFixed(1)}%</span>`;
}
function cColour(c){return COLOURS[Math.max(0,DATA.centres.indexOf(c))%COLOURS.length];}
function short(c){const t=DATA.centre_targets[c];return t?t.short:c.split(',')[0];}
function filterCentre(c,el){
  activeFilter=c;
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  el.classList.add('active');renderAll();
}
function filteredCentres(){return activeFilter==='ALL'?DATA.centres:[activeFilter];}
function filteredWeekly(){return activeFilter==='ALL'?DATA.weekly:DATA.weekly.filter(r=>r['Collection centre']===activeFilter);}
function initPage(){
  document.getElementById('weekBadge').textContent=DATA.week_label;
  document.getElementById('genDate').textContent=DATA.generated;
  document.getElementById('centreList').innerHTML=DATA.centres.join(' \u00a0|\u00a0 ');
  document.getElementById('topNote').innerHTML=
    `\u26a0\ufe0f <strong>Tests = individual investigations</strong> (each comma-separated item in the Investigations field is counted). ` +
    `<strong>Cases</strong> = number of case rows. Revenue = Total Fee (billed). Unique Referrers = distinct doctors/sources. ` +
    `<strong>${DATA.quarter}</strong> daily targets applied. Cancelled (${DATA.summary.cancelled}) excluded.`;
  const bar=document.getElementById('filterBar');
  DATA.centres.forEach((c,i)=>{
    const d=document.createElement('div');
    d.className='tab';
    d.innerHTML=`<span class="dot" style="background:${COLOURS[i%COLOURS.length]}"></span>${short(c)}`;
    d.onclick=()=>filterCentre(c,d);bar.appendChild(d);
  });
}
function renderKPI(){
  let inv=0,rev=0,paid=0,due=0,patients=0,refs=0;
  DATA.daily.forEach(r=>{
    if(activeFilter==='ALL'||r['Collection centre']===activeFilter){
      inv+=r.investigations||0;rev+=r.revenue||0;paid+=r.paid||0;due+=r.due||0;
    }
  });
  filteredWeekly().forEach(r=>{patients+=r.patients||0;refs+=r.unique_referrers||0;});
  document.getElementById('kpi-inv').textContent=fmtN(inv);
  document.getElementById('kpi-patients').textContent=fmtN(patients);
  document.getElementById('kpi-referrers').textContent=fmtN(refs);
  document.getElementById('kpi-revenue').textContent=fmt(rev);
  document.getElementById('kpi-paid').textContent=fmt(paid);
  document.getElementById('kpi-due').textContent=fmt(due);
  document.getElementById('kpi-rate').textContent=rev>0?(paid/rev*100).toFixed(1)+'%':'\u2013';
  const cents=filteredCentres();
  let tot=cents.reduce((s,c)=>s+(DATA.centre_targets[c]?.weekly||0),0);
  const sub=document.getElementById('kpi-rev-sub');
  sub.textContent=tot>0?`${(rev/tot*100).toFixed(1)}% of ${fmt(tot)} target`:'No target set';
}
function destroyChart(id){if(charts[id]){charts[id].destroy();delete charts[id];}}
function renderDailyInv(){
  destroyChart('di');
  const cents=filteredCentres();
  charts.di=new Chart(document.getElementById('chartDailyInv'),{
    type:'bar',
    data:{labels:DATA.date_labels,datasets:cents.map(c=>({
      label:short(c),backgroundColor:cColour(c)+'BB',borderColor:cColour(c),borderWidth:1,
      data:DATA.dates.map(d=>{const r=DATA.daily.find(x=>x.Date===d&&x['Collection centre']===c);return r?r.investigations:0;})
    }))},
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
      scales:{x:{stacked:false},y:{beginAtZero:true,ticks:{font:{size:10}}}}}
  });
}
function renderDailyRev(){
  destroyChart('dr');
  const cents=filteredCentres();
  charts.dr=new Chart(document.getElementById('chartDailyRev'),{
    type:'bar',
    data:{labels:DATA.date_labels,datasets:cents.map(c=>({
      label:short(c),backgroundColor:cColour(c)+'BB',borderColor:cColour(c),borderWidth:1,
      data:DATA.dates.map(d=>{const r=DATA.daily.find(x=>x.Date===d&&x['Collection centre']===c);return r?r.revenue:0;})
    }))},
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
      scales:{y:{beginAtZero:true,ticks:{callback:v=>'\u20a6'+v.toLocaleString(),font:{size:9}}}}}
  });
}
function renderVsTarget(){
  destroyChart('vt');
  const cents=filteredCentres();
  charts.vt=new Chart(document.getElementById('chartVsTarget'),{
    type:'bar',
    data:{
      labels:cents.map(c=>short(c)),
      datasets:[
        {label:'Actual Revenue',data:cents.map(c=>{const w=DATA.weekly.find(r=>r['Collection centre']===c);return w?w.revenue:0;}),backgroundColor:'#2E86ABBB',borderColor:'#2E86AB',borderWidth:1},
        {label:'Weekly Target',data:cents.map(c=>DATA.centre_targets[c]?.weekly||null),backgroundColor:'#E74C3C55',borderColor:'#E74C3C',borderWidth:2},
      ]
    },
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
      scales:{y:{beginAtZero:true,ticks:{callback:v=>'\u20a6'+v.toLocaleString(),font:{size:9}}}}}
  });
}
function renderCaseTypes(){
  destroyChart('ct');
  const cents=filteredCentres();
  const types=[...new Set(DATA.case_types.map(r=>r['Case Type']))];
  const tCols=['#2E86AB','#8E44AD','#16A085','#E67E22','#E74C3C','#F39C12','#27AE60'];
  charts.ct=new Chart(document.getElementById('chartCaseTypes'),{
    type:'bar',
    data:{labels:cents.map(c=>short(c)),datasets:types.map((t,i)=>({
      label:t,backgroundColor:tCols[i%tCols.length]+'BB',borderColor:tCols[i%tCols.length],borderWidth:1,
      data:cents.map(c=>DATA.case_types.filter(r=>r['Collection centre']===c&&r['Case Type']===t).reduce((s,r)=>s+r.investigations,0))
    }))},
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
      scales:{x:{stacked:true},y:{stacked:true,beginAtZero:true,ticks:{font:{size:10}}}}}
  });
}
function renderWeeklyRev(){
  destroyChart('wr');
  const cents=filteredCentres();
  charts.wr=new Chart(document.getElementById('chartWeeklyRev'),{
    type:'bar',
    data:{labels:cents.map(c=>short(c)),datasets:[
      {label:'Total Billed',data:cents.map(c=>{const w=DATA.weekly.find(r=>r['Collection centre']===c);return w?w.revenue:0;}),backgroundColor:'#2E86ABBB',borderColor:'#2E86AB',borderWidth:1},
      {label:'Collected',data:cents.map(c=>{const w=DATA.weekly.find(r=>r['Collection centre']===c);return w?w.paid:0;}),backgroundColor:'#27AE60BB',borderColor:'#27AE60',borderWidth:1},
    ]},
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
      scales:{y:{beginAtZero:true,ticks:{callback:v=>'\u20a6'+v.toLocaleString(),font:{size:9}}}}}
  });
}
function renderCollRate(){
  destroyChart('cr');
  const cents=filteredCentres();
  charts.cr=new Chart(document.getElementById('chartCollRate'),{
    type:'line',
    data:{labels:DATA.date_labels,datasets:cents.map(c=>({
      label:short(c),borderColor:cColour(c),backgroundColor:cColour(c)+'33',
      tension:.3,fill:false,pointRadius:4,borderWidth:2,
      data:DATA.dates.map(d=>{const r=DATA.daily.find(x=>x.Date===d&&x['Collection centre']===c);return(r&&r.revenue)?parseFloat((r.paid/r.revenue*100).toFixed(1)):null;})
    }))},
    options:{responsive:true,maintainAspectRatio:false,spanGaps:true,
      plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
      scales:{y:{min:0,max:105,ticks:{callback:v=>v+'%',font:{size:10}}}}}
  });
}
function renderDailyTable(){
  const rows=DATA.daily.filter(r=>activeFilter==='ALL'||r['Collection centre']===activeFilter)
    .sort((a,b)=>a.Date.localeCompare(b.Date)||a['Collection centre'].localeCompare(b['Collection centre']));
  document.getElementById('dailyTbody').innerHTML=rows.map(r=>{
    const col=cColour(r['Collection centre']);
    const sh=short(r['Collection centre']);
    const dt=DATA.centre_targets[r['Collection centre']]?.daily;
    const rev=r.revenue||0;
    const label=DATA.date_labels[DATA.dates.indexOf(r.Date)]||r.Date;
    const cr=rev>0?(r.paid/rev*100).toFixed(1)+'%':'\u2013';
    return`<tr>
      <td>${label}</td>
      <td><span class="dot" style="background:${col}"></span>${sh}</td>
      <td>${fmtN(r.investigations)}</td><td>${fmtN(r.cases)}</td><td>${fmtN(r.patients)}</td>
      <td>${fmtN(r.unique_referrers)}</td>
      <td>${fmt(r.revenue)}</td><td>${fmt(r.paid)}</td><td>${fmt(r.due)}</td>
      <td>${dt?fmt(dt):'<span class="badge badge-gray">No target</span>'}</td>
      <td>${achvBadge(rev,dt)}</td><td>${badge(cr,95,80)}</td>
    </tr>`;
  }).join('');
}
function renderWeeklyTable(){
  document.getElementById('weeklyTbody').innerHTML=filteredWeekly().map(r=>{
    const col=cColour(r['Collection centre']);
    const wt=DATA.centre_targets[r['Collection centre']]?.weekly;
    const rev=r.revenue||0;
    const cr=rev>0?(r.paid/rev*100).toFixed(1)+'%':'\u2013';
    const avg=r.cases>0?fmt(Math.round(rev/r.cases)):'\u2013';
    return`<tr>
      <td><span class="dot" style="background:${col}"></span>${r['Collection centre']}</td>
      <td>${fmtN(r.investigations)}</td><td>${fmtN(r.cases)}</td><td>${fmtN(r.patients)}</td>
      <td>${fmtN(r.unique_referrers)}</td>
      <td>${fmt(r.revenue)}</td><td>${fmt(r.paid)}</td><td>${fmt(r.due)}</td>
      <td>${wt?fmt(wt):'<span class="badge badge-gray">No target</span>'}</td>
      <td>${achvBadge(rev,wt)}</td><td>${avg}</td><td>${badge(cr,95,80)}</td>
    </tr>`;
  }).join('');
}
function renderRefTable(){
  const centreTotal={};
  DATA.weekly.forEach(r=>{centreTotal[r['Collection centre']]=r.cases;});
  const cents=filteredCentres();
  const rows=DATA.top_referrers.filter(r=>cents.includes(r['Collection centre']));
  document.getElementById('refTbody').innerHTML=rows.map((r,i)=>{
    const col=cColour(r['Collection centre']);
    const tot=centreTotal[r['Collection centre']]||1;
    return`<tr>
      <td>${i+1}</td>
      <td><span class="dot" style="background:${col}"></span>${short(r['Collection centre'])}</td>
      <td>${r.Referrer}</td><td>${r.cases}</td>
      <td>${fmt(r.revenue)}</td><td>${(r.cases/tot*100).toFixed(1)}%</td>
    </tr>`;
  }).join('');
}
function renderAgentTable(){
  let rows;
  if(activeFilter!=='ALL'){
    rows=DATA.agent_centre.filter(r=>r['Collection centre']===activeFilter)
      .sort((a,b)=>b.investigations-a.investigations)
      .map((r,i)=>{
        const cr=r.revenue>0?(r.paid/r.revenue*100).toFixed(1)+'%':'\u2013';
        return`<tr><td>${i+1}</td><td>${r.Agent}</td><td>${fmtN(r.investigations)}</td>
          <td>${fmtN(r.cases)}</td><td>\u2013</td>
          <td>${fmt(r.revenue)}</td><td>${fmt(r.paid)}</td><td>${badge(cr,95,80)}</td></tr>`;
      });
  }else{
    rows=DATA.agent_total.map((r,i)=>{
      const cr=r.revenue>0?(r.paid/r.revenue*100).toFixed(1)+'%':'\u2013';
      return`<tr><td>${i+1}</td><td>${r.Agent}</td><td>${fmtN(r.investigations)}</td>
        <td>${fmtN(r.cases)}</td><td>${fmtN(r.patients)}</td>
        <td>${fmt(r.revenue)}</td><td>${fmt(r.paid)}</td><td>${badge(cr,95,80)}</td></tr>`;
    });
  }
  document.getElementById('agentTbody').innerHTML=rows.join('');
}
function renderAll(){
  renderKPI();
  renderDailyInv();renderDailyRev();renderVsTarget();renderCaseTypes();
  renderWeeklyRev();renderCollRate();
  renderDailyTable();renderWeeklyTable();renderRefTable();renderAgentTable();
}
document.addEventListener('DOMContentLoaded',()=>{initPage();renderAll();});
</script></body></html>
""")

    return ''.join(html_parts)


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        csvs = [f for f in os.listdir(WORKSPACE) if f.lower().endswith('.csv') and 'cases' in f.lower()]
        if not csvs:
            print("Usage: python3 update_dashboard.py <path_to_csv>"); sys.exit(1)
        csvs.sort(key=lambda f: os.path.getmtime(os.path.join(WORKSPACE, f)), reverse=True)
        csv_path = os.path.join(WORKSPACE, csvs[0])
        print(f"Auto-detected: {csv_path}")
    else:
        csv_path = sys.argv[1]

    if not os.path.exists(csv_path):
        print(f"ERROR: File not found: {csv_path}"); sys.exit(1)

    data, df_active, df_full, week_label = process_csv(csv_path)

    print(f"\n{'='*55}")
    print(f"Week:                 {week_label}")
    print(f"Quarter:              {data['quarter']}")
    print(f"Total Investigations: {data['summary']['total_investigations']}")
    print(f"Total Case Rows:      {data['summary']['total_cases']}")
    print(f"Unique Patients:      {data['summary']['total_patients']}")
    print(f"Unique Referrers:     {data['summary']['total_unique_referrers']}")
    print(f"Total Revenue:        \u20a6{data['summary']['total_revenue']:,}")
    print(f"Collection Rate:      {data['summary']['collection_rate']}%")
    print(f"Cancelled:            {data['summary']['cancelled']}")
    print(f"{'='*55}\n")

    xlsx_path = build_excel(data, df_active, week_label)

    # Single-week HTML (v2)
    html_content = build_html(data)
    html_path = os.path.join(WORKSPACE, 'Sure_Weekly_Dashboard.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    dated = os.path.join(WORKSPACE, f"Sure_Weekly_Dashboard_{datetime.date.today().strftime('%Y-%m-%d')}.html")
    shutil.copy2(html_path, dated)
    print(f"HTML (single-week) saved: {html_path}")

    # History + multi-period dashboard (v3)
    history = load_history()
    save_to_history(history, data)
    multi_html_path = build_html_v3(history)

    # Copy all outputs to permanent workspace
    perm = _perm_workspace()
    xl_dest    = os.path.join(perm, os.path.basename(xlsx_path))
    html_dest  = os.path.join(perm, 'Sure_Weekly_Dashboard.html')
    multi_dest = os.path.join(perm, 'Sure_Weekly_Dashboard_Multi.html')
    for src, dst in [(xlsx_path, xl_dest), (html_path, html_dest), (multi_html_path, multi_dest)]:
        if os.path.abspath(src) != os.path.abspath(dst):
            shutil.copy2(src, dst)

    with open(os.path.join(WORKSPACE, 'dashboard_data.json'), 'w') as f:
        json.dump(data, f, indent=2, default=str)

    # Update live Artifacts dashboards (sidebar pins)
    update_artifact_html(data)

    print("\n\u2705 Dashboard update complete!")
    print(f"   Excel        \u2192 {xl_dest}")
    print(f"   Weekly view  \u2192 {html_dest}")
    print(f"   Multi-period \u2192 {multi_dest}")

# ─────────────────────────────────────────────────────────────────────────────
# ARTIFACT HTML SYNC
# Injects the new week into the live Cowork sidebar dashboards so they stay
# current without any manual copy step.
# ─────────────────────────────────────────────────────────────────────────────

ARTIFACT_HTML_PATHS = [
    '/Users/mac/Documents/Claude/Artifacts/sure-weekly-dashboard/index.html',
    '/Users/mac/Documents/01_Sure_Diagnostics/Dashboards/Sure Labsmart Dashboard/weekly/index.html',
]

def update_artifact_html(data):
    """Append/update this week's entry in all Artifacts-format HTML dashboards."""
    import re as _re

    week_start = data['dates'][0]          # e.g. '2026-04-13'
    ctargs     = data['centre_targets']    # {name: {short, color, weekly, ...}}

    # Build per-centre list in Artifacts schema
    weekly_by_centre = {r['Collection centre']: r for r in data['weekly']}
    centres_list = []
    for name in sorted(weekly_by_centre.keys()):
        r   = weekly_by_centre[name]
        tgt = ctargs.get(name, {})
        weekly_tgt = tgt.get('weekly') or 0
        rev = int(r['revenue'])
        att = round(rev / weekly_tgt * 100, 1) if weekly_tgt else 0
        centres_list.append({
            'name':          name,
            'short':         tgt.get('short', name.split(',')[0]),
            'color':         tgt.get('color', '#888888'),
            'revenue':       rev,
            'weekly_target': int(weekly_tgt),
            'attainment':    att,
            'cases':         int(r['cases']),
            'patients':      int(r['patients']),
            'investigations':int(r['investigations']),
            'referrers':     int(r['unique_referrers']),
        })

    month_int   = int(week_start[5:7])
    year_int    = int(week_start[:4])
    quarter_int = (month_int - 1) // 3 + 1

    new_entry = {
        'week_key':   week_start,
        'label':      data['week_label'],
        'year':       year_int,
        'month':      month_int,
        'quarter':    f'Q{quarter_int} {year_int}',
        'summary':    data['summary'],
        'centres':    centres_list,
        'referrers':  data['top_referrers'],
        'agents':     data['agent_total'],
    }

    pattern = _re.compile(r'const HISTORY = (\[.*?\]);', _re.DOTALL)

    for path in ARTIFACT_HTML_PATHS:
        if not os.path.exists(path):
            print(f"   ⚠ Artifact not found, skipping: {path}")
            continue
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()
        m = pattern.search(content)
        if not m:
            print(f"   ⚠ HISTORY constant not found in: {path}")
            continue
        history = json.loads(m.group(1))
        idx = next((i for i, e in enumerate(history) if e.get('week_key') == week_start), None)
        if idx is not None:
            history[idx] = new_entry
        else:
            history.append(new_entry)
        new_content = (
            content[:m.start()]
            + 'const HISTORY = ' + json.dumps(history, separators=(',', ':')) + ';'
            + content[m.end():]
        )
        with open(path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        print(f"   Artifact  → {path}")


# ─────────────────────────────────────────────────────────────────────────────
# HISTORY MANAGEMENT (v3)
# ─────────────────────────────────────────────────────────────────────────────

def _perm_workspace():
    """Return the permanent workspace directory (survives sessions)."""
    mnt = os.path.join(WORKSPACE, 'mnt', 'Sure Finance & Business Scorecards')
    return mnt if os.path.isdir(mnt) else WORKSPACE

HISTORY_FILE = None  # resolved lazily

def _history_file():
    global HISTORY_FILE
    if HISTORY_FILE is None:
        HISTORY_FILE = os.path.join(_perm_workspace(), 'dashboard_history.json')
    return HISTORY_FILE

def load_history():
    """Load accumulated weekly history; return empty dict if none."""
    hf = _history_file()
    if os.path.exists(hf):
        with open(hf, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'weeks': {}}

def save_to_history(history, data):
    """Append/overwrite a week's data in history and persist to disk."""
    week_start = data['dates'][0] if data.get('dates') else 'unknown'
    history['weeks'][week_start] = {
        'week_label':     data['week_label'],
        'year':           int(week_start[:4]),
        'month':          int(week_start[5:7]),
        'quarter':        data['quarter'],
        'num_days':       data.get('num_days', 7),
        'centres':        data['centres'],
        'dates':          data['dates'],
        'date_labels':    data.get('date_labels', []),
        'summary':        data['summary'],
        'daily':          data['daily'],
        'weekly':         data['weekly'],
        'top_referrers':  data.get('top_referrers', []),
        'agent_total':    data.get('agent_total', []),
        'centre_targets': data.get('centre_targets', {}),
    }
    hf = _history_file()
    with open(hf, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, default=str)
    print(f"History updated → {hf}  ({len(history['weeks'])} weeks stored)")
    return history


# ─────────────────────────────────────────────────────────────────────────────
# MULTI-PERIOD HTML DASHBOARD (v3)
# ─────────────────────────────────────────────────────────────────────────────

def build_html_v3(history):
    """Build multi-period dashboard with year/month/week filters + summary/drill-down toggle."""

    history_json  = json.dumps(history,  indent=2, default=str)
    targets_json  = json.dumps(TARGETS,  indent=2, default=str)
    generated     = datetime.date.today().strftime('%d %b %Y')

    p = []  # html parts

    # ── HEAD ──────────────────────────────────────────────────────────────────
    p.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Sure Finance \u2013 Performance Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root{{--navy:#1F3864;--teal:#2E86AB;--green:#27AE60;--amber:#E67E22;--red:#E74C3C;
      --purple:#8E44AD;--dgreen:#16A085;--gold:#F39C12;--bg:#F0F4F8;--card:#fff;--border:#DDE3EA;}}
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:Arial,sans-serif;background:var(--bg);color:#2C3E50;font-size:13px;}}
.header{{background:var(--navy);color:#fff;padding:14px 24px;display:flex;justify-content:space-between;align-items:center;}}
.header h1{{font-size:16px;font-weight:700;letter-spacing:.4px;}}
.header .meta{{font-size:11px;opacity:.75;text-align:right;}}
.ctrl-bar{{background:#fff;border-bottom:2px solid var(--border);padding:10px 24px;
  display:flex;align-items:center;gap:10px;flex-wrap:wrap;}}
.ctrl-bar label{{font-size:11px;font-weight:700;color:#666;margin-right:2px;}}
.ctrl-bar select{{font-size:12px;padding:5px 8px;border:1px solid var(--border);
  border-radius:6px;background:#fff;cursor:pointer;color:var(--navy);font-weight:600;}}
.ctrl-bar select:focus{{outline:none;border-color:var(--teal);}}
.mode-wrap{{margin-left:auto;display:flex;gap:6px;}}
.mode-btn{{padding:6px 16px;border:2px solid var(--border);border-radius:20px;
  font-size:11px;font-weight:700;cursor:pointer;background:#fff;color:#666;transition:.2s;}}
.mode-btn.active{{background:var(--teal);border-color:var(--teal);color:#fff;}}
.centre-bar{{background:#fff;border-bottom:1px solid var(--border);padding:0 24px;
  display:flex;gap:0;overflow-x:auto;}}
.tab{{padding:9px 16px;cursor:pointer;font-size:12px;font-weight:600;color:#666;
  border-bottom:3px solid transparent;transition:.2s;white-space:nowrap;}}
.tab:hover{{color:var(--navy);background:#F8FAFC;}}
.tab.active{{color:var(--navy);border-bottom-color:var(--teal);}}
.period-label{{background:linear-gradient(135deg,var(--teal),var(--navy));color:#fff;
  padding:7px 24px;font-size:12px;font-weight:700;letter-spacing:.4px;}}
.main{{padding:18px 24px;max-width:1600px;margin:0 auto;}}
.section-title{{font-size:12px;font-weight:700;color:var(--navy);text-transform:uppercase;
  letter-spacing:.8px;margin:18px 0 10px;padding-bottom:5px;border-bottom:2px solid var(--border);}}
.kpi-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:10px;margin-bottom:18px;}}
.kpi-card{{background:var(--card);border-radius:10px;padding:13px 14px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);border-top:4px solid var(--teal);text-align:center;}}
.kpi-label{{font-size:10px;text-transform:uppercase;letter-spacing:.5px;color:#888;font-weight:700;}}
.kpi-value{{font-size:21px;font-weight:800;color:var(--navy);margin:5px 0 2px;}}
.kpi-sub{{font-size:10px;color:#aaa;}}
.charts-grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px;}}
.charts-grid-3{{display:grid;grid-template-columns:2fr 1fr;gap:14px;}}
.chart-card{{background:var(--card);border-radius:10px;padding:14px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);}}
.chart-title{{font-size:12px;font-weight:700;color:var(--navy);margin-bottom:10px;}}
.chart-wrap{{position:relative;height:210px;}}
.chart-wrap-tall{{position:relative;height:260px;}}
.table-card{{background:var(--card);border-radius:10px;padding:14px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:14px;overflow-x:auto;}}
table{{width:100%;border-collapse:collapse;}}
thead tr{{background:var(--navy);color:#fff;}}
thead th{{padding:7px 9px;font-size:11px;font-weight:700;text-align:center;white-space:nowrap;}}
thead th:first-child{{text-align:left;}}
tbody tr:nth-child(even){{background:#F7FAFC;}}
tbody tr:hover{{background:#EBF5FB;}}
tbody td{{padding:6px 9px;border-bottom:1px solid var(--border);font-size:12px;
  text-align:center;white-space:nowrap;}}
tbody td:first-child{{text-align:left;font-weight:600;}}
.badge{{display:inline-block;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:700;}}
.badge-green{{background:#D5F5E3;color:#1E8449;}}
.badge-amber{{background:#FCF3CF;color:#9A7D0A;}}
.badge-red{{background:#FADBD8;color:#922B21;}}
.badge-gray{{background:#EAEDED;color:#555;}}
.dot{{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:4px;vertical-align:middle;}}
.note{{background:#FEF9E7;border-left:4px solid var(--gold);padding:7px 12px;
  font-size:11px;color:#7D6608;border-radius:0 6px 6px 0;margin-bottom:14px;}}
.tgt-row{{background:#E8F4FD!important;font-weight:700!important;}}
.total-row td{{background:var(--teal)!important;color:#fff!important;font-weight:700;}}
.footer{{text-align:center;padding:14px;font-size:10px;color:#aaa;background:#fff;
  border-top:1px solid var(--border);margin-top:20px;}}
.hidden{{display:none;}}
@media(max-width:900px){{
  .kpi-grid{{grid-template-columns:repeat(3,1fr);}}
  .charts-grid,.charts-grid-3{{grid-template-columns:1fr;}}
}}
</style></head><body>
""")

    # ── HEADER ────────────────────────────────────────────────────────────────
    p.append(f"""
<div class="header">
  <div>
    <h1>&#128202; SURE FINANCE &amp; BUSINESS \u2014 PERFORMANCE DASHBOARD</h1>
    <div style="font-size:11px;opacity:.7;margin-top:3px" id="centreList"></div>
  </div>
  <div class="meta">
    <div>Generated: {generated}</div>
    <div style="margin-top:4px" id="weekCount"></div>
  </div>
</div>
""")

    # ── CONTROLS ──────────────────────────────────────────────────────────────
    p.append("""
<div class="ctrl-bar">
  <label>Year</label>
  <select id="yrSel" onchange="onYearChange(+this.value)"></select>
  <label style="margin-left:8px">Month</label>
  <select id="monSel" onchange="onMonthChange(+this.value)"></select>
  <label style="margin-left:8px">Week</label>
  <select id="wkSel" onchange="onWeekChange(this.value)"></select>
  <div class="mode-wrap">
    <button class="mode-btn active" id="btnSum" onclick="setMode('summary')">&#128202; Summary</button>
    <button class="mode-btn" id="btnDrill" onclick="setMode('drill')">&#128269; Weekly View</button>
  </div>
</div>
<div class="centre-bar" id="centreBar">
  <div class="tab active" onclick="filterCentre('ALL',this)">&#127970; All Centres</div>
</div>
<div class="period-label" id="periodLabel"></div>
""")

    # ── MAIN CONTENT ──────────────────────────────────────────────────────────
    p.append("""
<div class="main">
<div class="note" id="topNote"></div>

<!-- ═══════════ SUMMARY VIEW ═══════════ -->
<div id="summaryView">
  <div class="section-title">Key Performance Indicators</div>
  <div class="kpi-grid">
    <div class="kpi-card" style="border-top-color:var(--teal)">
      <div class="kpi-label">Total Tests</div>
      <div class="kpi-value" id="kpi-inv">\u2013</div>
      <div class="kpi-sub">Individual investigations</div></div>
    <div class="kpi-card" style="border-top-color:var(--dgreen)">
      <div class="kpi-label">Cases</div>
      <div class="kpi-value" id="kpi-cases">\u2013</div>
      <div class="kpi-sub">Case rows</div></div>
    <div class="kpi-card" style="border-top-color:var(--purple)">
      <div class="kpi-label">Unique Referrers</div>
      <div class="kpi-value" id="kpi-refs">\u2013</div>
      <div class="kpi-sub">Distinct doctors/sources</div></div>
    <div class="kpi-card" style="border-top-color:var(--navy)">
      <div class="kpi-label">Total Revenue</div>
      <div class="kpi-value" id="kpi-rev">\u20a6\u2013</div>
      <div class="kpi-sub" id="kpi-rev-sub">vs target</div></div>
    <div class="kpi-card" style="border-top-color:var(--green)">
      <div class="kpi-label">Fee Collected</div>
      <div class="kpi-value" id="kpi-paid">\u20a6\u2013</div>
      <div class="kpi-sub">Amount paid</div></div>
    <div class="kpi-card" style="border-top-color:var(--amber)">
      <div class="kpi-label">Fee Due</div>
      <div class="kpi-value" id="kpi-due">\u20a6\u2013</div>
      <div class="kpi-sub">Outstanding balance</div></div>
    <div class="kpi-card" style="border-top-color:var(--red)">
      <div class="kpi-label">Collection Rate</div>
      <div class="kpi-value" id="kpi-rate">\u2013</div>
      <div class="kpi-sub">Paid / Total fee</div></div>
  </div>

  <div class="section-title">Revenue &amp; Performance Trends</div>
  <div class="charts-grid">
    <div class="chart-card">
      <div class="chart-title" id="trendTitle">&#128176; Revenue by Week (\u20a6)</div>
      <div class="chart-wrap-tall"><canvas id="chartTrend"></canvas></div></div>
    <div class="chart-card">
      <div class="chart-title">&#127919; Revenue vs Target by Centre (\u20a6)</div>
      <div class="chart-wrap-tall"><canvas id="chartVsTarget"></canvas></div></div>
  </div>

  <div class="charts-grid" style="margin-top:14px">
    <div class="chart-card">
      <div class="chart-title" id="invTrendTitle">&#128203; Investigations by Week</div>
      <div class="chart-wrap"><canvas id="chartInvTrend"></canvas></div></div>
    <div class="chart-card">
      <div class="chart-title">&#128179; Revenue vs Collected by Centre (\u20a6)</div>
      <div class="chart-wrap"><canvas id="chartRevColl"></canvas></div></div>
  </div>

  <div class="section-title">Centre Performance Summary</div>
  <div class="table-card"><table>
  <thead><tr>
    <th>Collection Centre</th><th>Investigations</th><th>Cases</th>
    <th>Total Revenue (\u20a6)</th><th>Collected (\u20a6)</th><th>Due (\u20a6)</th>
    <th>Period Target (\u20a6)</th><th>Achievement</th><th>Collection %</th>
    <th>Avg Fee/Case (\u20a6)</th>
  </tr></thead>
  <tbody id="centreTbody"></tbody></table></div>

  <div class="section-title">Top Referrers</div>
  <div class="table-card"><table>
  <thead><tr>
    <th>#</th><th>Centre</th><th>Referrer / Doctor</th>
    <th>Cases</th><th>Revenue (\u20a6)</th><th>% of Centre Rev</th>
  </tr></thead>
  <tbody id="refTbody"></tbody></table></div>

  <div class="section-title">Agent Performance</div>
  <div class="table-card"><table>
  <thead><tr>
    <th>#</th><th>Agent</th><th>Investigations</th><th>Cases</th>
    <th>Revenue (\u20a6)</th><th>Collected (\u20a6)</th><th>Collection %</th>
  </tr></thead>
  <tbody id="agentTbody"></tbody></table></div>
</div><!-- end summaryView -->

<!-- ═══════════ DRILL-DOWN VIEW ═══════════ -->
<div id="drillView" class="hidden">
  <div id="drillNote" class="note"></div>
  <div class="section-title">Key Performance Indicators \u2014 Selected Week</div>
  <div class="kpi-grid">
    <div class="kpi-card" style="border-top-color:var(--teal)">
      <div class="kpi-label">Total Tests</div>
      <div class="kpi-value" id="dkpi-inv">\u2013</div>
      <div class="kpi-sub">Investigations</div></div>
    <div class="kpi-card" style="border-top-color:var(--dgreen)">
      <div class="kpi-label">Cases</div>
      <div class="kpi-value" id="dkpi-cases">\u2013</div>
      <div class="kpi-sub">Case rows</div></div>
    <div class="kpi-card" style="border-top-color:var(--purple)">
      <div class="kpi-label">Unique Referrers</div>
      <div class="kpi-value" id="dkpi-refs">\u2013</div>
      <div class="kpi-sub">Distinct sources</div></div>
    <div class="kpi-card" style="border-top-color:var(--navy)">
      <div class="kpi-label">Total Revenue</div>
      <div class="kpi-value" id="dkpi-rev">\u20a6\u2013</div>
      <div class="kpi-sub" id="dkpi-rev-sub">vs weekly target</div></div>
    <div class="kpi-card" style="border-top-color:var(--green)">
      <div class="kpi-label">Fee Collected</div>
      <div class="kpi-value" id="dkpi-paid">\u20a6\u2013</div>
      <div class="kpi-sub">Amount paid</div></div>
    <div class="kpi-card" style="border-top-color:var(--amber)">
      <div class="kpi-label">Fee Due</div>
      <div class="kpi-value" id="dkpi-due">\u20a6\u2013</div>
      <div class="kpi-sub">Outstanding</div></div>
    <div class="kpi-card" style="border-top-color:var(--red)">
      <div class="kpi-label">Collection Rate</div>
      <div class="kpi-value" id="dkpi-rate">\u2013</div>
      <div class="kpi-sub">Paid / Total fee</div></div>
  </div>

  <div class="section-title">Daily Breakdown Charts</div>
  <div class="charts-grid">
    <div class="chart-card">
      <div class="chart-title">&#128197; Daily Investigations by Centre</div>
      <div class="chart-wrap"><canvas id="dChartDailyInv"></canvas></div></div>
    <div class="chart-card">
      <div class="chart-title">&#128176; Daily Revenue by Centre (\u20a6)</div>
      <div class="chart-wrap"><canvas id="dChartDailyRev"></canvas></div></div>
  </div>
  <div class="charts-grid" style="margin-top:14px">
    <div class="chart-card">
      <div class="chart-title">&#127919; Revenue vs Weekly Target (\u20a6)</div>
      <div class="chart-wrap"><canvas id="dChartVsTarget"></canvas></div></div>
    <div class="chart-card">
      <div class="chart-title">&#128200; Daily Collection Rate (%)</div>
      <div class="chart-wrap"><canvas id="dChartCollRate"></canvas></div></div>
  </div>

  <div class="section-title">Daily Performance Summary</div>
  <div class="table-card"><table>
  <thead><tr>
    <th>Date</th><th>Centre</th><th>Investigations</th><th>Cases</th><th>Patients</th>
    <th>Unique Referrers</th><th>Total Fee (\u20a6)</th><th>Paid (\u20a6)</th><th>Due (\u20a6)</th>
    <th>Daily Target (\u20a6)</th><th>Achievement</th><th>Collection %</th>
  </tr></thead>
  <tbody id="dailyTbody"></tbody></table></div>

  <div class="section-title">Weekly Centre Summary vs Target</div>
  <div class="table-card"><table>
  <thead><tr>
    <th>Collection Centre</th><th>Investigations</th><th>Cases</th><th>Patients</th>
    <th>Unique Referrers</th><th>Total Fee (\u20a6)</th><th>Paid (\u20a6)</th><th>Due (\u20a6)</th>
    <th>Weekly Target (\u20a6)</th><th>Achievement</th><th>Collection %</th>
  </tr></thead>
  <tbody id="weeklyTbody"></tbody></table></div>

  <div class="section-title">Top Referrers</div>
  <div class="table-card"><table>
  <thead><tr>
    <th>#</th><th>Centre</th><th>Referrer / Doctor</th>
    <th>Cases</th><th>Revenue (\u20a6)</th>
  </tr></thead>
  <tbody id="dRefTbody"></tbody></table></div>

  <div class="section-title">Agent Performance</div>
  <div class="table-card"><table>
  <thead><tr>
    <th>#</th><th>Agent</th><th>Investigations</th><th>Cases</th>
    <th>Revenue (\u20a6)</th><th>Collected (\u20a6)</th><th>Collection %</th>
  </tr></thead>
  <tbody id="dAgentTbody"></tbody></table></div>
</div><!-- end drillView -->
</div><!-- end main -->
<div class="footer">Sure Finance &amp; Business Scorecards &nbsp;|&nbsp; Auto-generated from LabSmartLIS &nbsp;|&nbsp; Multi-period v3</div>
""")

    # ── JAVASCRIPT ────────────────────────────────────────────────────────────
    p.append('<script>\nconst HISTORY=')
    p.append(history_json)
    p.append(';\nconst TARGETS_CFG=')
    p.append(targets_json)
    p.append(""";
const COLOURS=['#2E86AB','#8E44AD','#16A085','#E67E22','#E74C3C','#F39C12'];
const MONTH_NAMES=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
const charts={};
let state={year:null,month:0,week:null,mode:'summary',centre:'ALL'};

// ── Formatting ──────────────────────────────────────────────────────────────
const fmt=n=>n==null?'\u2013':'\u20a6'+Math.round(n).toLocaleString();
const fmtN=n=>n==null?'\u2013':Number(n).toLocaleString();
function badge(val,hi=100,lo=80){
  if(val==null||val==='\u2013')return`<span class="badge badge-gray">\u2013</span>`;
  const n=parseFloat(val);
  if(isNaN(n))return`<span class="badge badge-gray">${val}</span>`;
  const cls=n>=hi?'badge-green':n>=lo?'badge-amber':'badge-red';
  return`<span class="badge ${cls}">${val}</span>`;
}
function achvBadge(actual,target){
  if(!target||!actual)return`<span class="badge badge-gray">${target?'\u2013':'No target'}</span>`;
  const p=actual/target*100;
  const cls=p>=100?'badge-green':p>=80?'badge-amber':'badge-red';
  return`<span class="badge ${cls}">${p.toFixed(1)}%</span>`;
}
function short(c){const t=TARGETS_CFG[c];return t?t.short:c.split(',')[0];}
function cCol(c){const cs=allCentres();return COLOURS[Math.max(0,cs.indexOf(c))%COLOURS.length];}

// ── Data helpers ─────────────────────────────────────────────────────────────
function allCentres(){
  const cs=new Set();
  Object.values(HISTORY.weeks).forEach(w=>(w.centres||[]).forEach(c=>cs.add(c)));
  return [...cs].sort();
}
function getYears(){
  return[...new Set(Object.keys(HISTORY.weeks).map(k=>+k.slice(0,4)))].sort((a,b)=>b-a);
}
function getMonths(year){
  const ms=[...new Set(Object.keys(HISTORY.weeks).filter(k=>+k.slice(0,4)===year).map(k=>+k.slice(5,7)))].sort((a,b)=>a-b);
  return[0,...ms]; // 0 = All months
}
function getWeekKeys(year,month){
  return Object.keys(HISTORY.weeks).filter(k=>{
    const y=+k.slice(0,4),m=+k.slice(5,7);
    return y===year&&(month===0||m===month);
  }).sort();
}

// Period target: sum of daily_target × 7 for each week in the period, per centre
function periodTarget(centre,weekKeys){
  if(!TARGETS_CFG[centre])return null;
  let total=0;
  for(const key of weekKeys){
    const m=+key.slice(5,7);
    const qi=Math.floor((m-1)/3);
    const d=(TARGETS_CFG[centre].daily||[])[qi]||0;
    total+=d*7;
  }
  return total||null;
}

// ── Aggregation ──────────────────────────────────────────────────────────────
function aggregate(weekKeys,filterCentre){
  const sum={investigations:0,cases:0,patients:0,referrers:0,revenue:0,paid:0,due:0,cancelled:0};
  const centreMap={};
  const trend=[];
  const refMap={};
  const agentMap={};

  for(const key of weekKeys){
    const w=HISTORY.weeks[key];
    if(!w)continue;
    const centres=filterCentre==='ALL'?(w.centres||[]):(w.centres||[]).filter(c=>c===filterCentre);
    const wlyCentres=(w.weekly||[]).filter(r=>filterCentre==='ALL'||r['Collection centre']===filterCentre);

    let wRev=0,wPaid=0,wDue=0,wInv=0,wCases=0,wPat=0,wRefs=0;
    for(const cr of wlyCentres){
      const c=cr['Collection centre'];
      if(!centreMap[c])centreMap[c]={centre:c,investigations:0,cases:0,revenue:0,paid:0,due:0,patients:0,referrers:0};
      centreMap[c].investigations+=(cr.investigations||0);
      centreMap[c].cases+=(cr.cases||0);
      centreMap[c].revenue+=(cr.revenue||0);
      centreMap[c].paid+=(cr.paid||0);
      centreMap[c].due+=(cr.due||0);
      centreMap[c].patients+=(cr.patients||0);
      centreMap[c].referrers+=(cr.unique_referrers||0);
      wRev+=(cr.revenue||0);wPaid+=(cr.paid||0);wDue+=(cr.due||0);
      wInv+=(cr.investigations||0);wCases+=(cr.cases||0);
      wPat+=(cr.patients||0);wRefs+=(cr.unique_referrers||0);
    }
    if(filterCentre==='ALL'){
      sum.patients+=(w.summary.total_patients||0);
      sum.referrers+=(w.summary.total_unique_referrers||0);
      sum.cancelled+=(w.summary.cancelled||0);
    } else {
      sum.patients+=wPat; sum.referrers+=wRefs;
    }
    sum.revenue+=wRev;sum.paid+=wPaid;sum.due+=wDue;
    sum.investigations+=wInv;sum.cases+=wCases;

    trend.push({key,label:w.week_label||key,revenue:wRev,paid:wPaid,investigations:wInv,cases:wCases});

    // Referrers
    for(const r of(w.top_referrers||[])){
      if(filterCentre!=='ALL'&&r['Collection centre']!==filterCentre)continue;
      const k2=r['Referrer']||'Unknown';
      if(!refMap[k2])refMap[k2]={ref:k2,centre:r['Collection centre'],cases:0,revenue:0};
      refMap[k2].cases+=(r.cases||0);refMap[k2].revenue+=(r.revenue||0);
    }
    // Agents
    for(const a of(w.agent_total||[])){
      const ak=a['Agent']||'Unknown';
      if(!agentMap[ak])agentMap[ak]={agent:ak,investigations:0,cases:0,revenue:0,paid:0};
      agentMap[ak].investigations+=(a.investigations||0);
      agentMap[ak].cases+=(a.cases||0);
      agentMap[ak].revenue+=(a.revenue||0);
      agentMap[ak].paid+=(a.paid||0);
    }
  }
  sum.collection_rate=sum.revenue>0?(sum.paid/sum.revenue*100).toFixed(1):0;
  return{
    summary:sum,
    centreBreakdown:Object.values(centreMap).sort((a,b)=>b.revenue-a.revenue),
    trend:trend.sort((a,b)=>a.key.localeCompare(b.key)),
    topReferrers:Object.values(refMap).sort((a,b)=>b.cases-a.cases).slice(0,40),
    agents:Object.values(agentMap).sort((a,b)=>b.investigations-a.investigations),
  };
}

// ── Chart helpers ────────────────────────────────────────────────────────────
function destroyChart(id){if(charts[id]){charts[id].destroy();delete charts[id];}}
function mkChart(id,config){destroyChart(id);charts[id]=new Chart(document.getElementById(id).getContext('2d'),config);}

// ── Filter controls ──────────────────────────────────────────────────────────
function filterCentre(c,el){
  state.centre=c;
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  el.classList.add('active');
  renderAll();
}
function setMode(m){
  state.mode=m;
  document.getElementById('btnSum').classList.toggle('active',m==='summary');
  document.getElementById('btnDrill').classList.toggle('active',m==='drill');
  document.getElementById('summaryView').classList.toggle('hidden',m==='drill');
  document.getElementById('drillView').classList.toggle('hidden',m==='summary');
  renderAll();
}
function onYearChange(y){
  state.year=y;state.month=0;state.week=null;
  populateMonths(y);populateWeeks(y,0);renderAll();
}
function onMonthChange(m){
  state.month=m;state.week=null;
  populateWeeks(state.year,m);renderAll();
}
function onWeekChange(v){state.week=v||null;renderAll();}

function populateYears(years){
  const s=document.getElementById('yrSel');s.innerHTML='';
  years.forEach(y=>{const o=document.createElement('option');o.value=y;o.text=y;s.appendChild(o);});
  s.value=state.year||years[0];
}
function populateMonths(year){
  const s=document.getElementById('monSel');s.innerHTML='';
  getMonths(year).forEach(m=>{
    const o=document.createElement('option');o.value=m;
    o.text=m===0?'All Months':MONTH_NAMES[m];s.appendChild(o);
  });
  s.value=state.month;
}
function populateWeeks(year,month){
  const s=document.getElementById('wkSel');s.innerHTML='';
  const allOpt=document.createElement('option');allOpt.value='';allOpt.text='All Weeks';s.appendChild(allOpt);
  getWeekKeys(year,month).forEach(key=>{
    const w=HISTORY.weeks[key];
    const o=document.createElement('option');o.value=key;
    o.text=w?w.week_label:key;s.appendChild(o);
  });
  s.value=state.week||'';
}

// ── Current week keys (what to show) ─────────────────────────────────────────
function currentWeekKeys(){
  if(state.week)return[state.week];
  return getWeekKeys(state.year,state.month);
}

// ── Period label ─────────────────────────────────────────────────────────────
function updatePeriodLabel(weekKeys){
  const numW=weekKeys.length;
  let label='';
  if(state.week){
    const w=HISTORY.weeks[state.week];
    label=w?'Week: '+w.week_label:'Week: '+state.week;
  } else if(state.month===0){
    label='Full Year '+state.year+' — '+numW+' week'+(numW!==1?'s':'')+' of data';
  } else {
    label=MONTH_NAMES[state.month]+' '+state.year+' — '+numW+' week'+(numW!==1?'s':'')+' of data';
  }
  const q=state.month?'Q'+Math.ceil(state.month/3)+' ':'';
  document.getElementById('periodLabel').textContent=`${label}   \u00b7   ${q}${state.year} targets applied`;
}

// ── Top note ─────────────────────────────────────────────────────────────────
function updateNote(weekKeys){
  const s=Object.keys(HISTORY.weeks).length;
  document.getElementById('topNote').innerHTML=
    `\u26A0\uFE0F Tests = individual investigations (comma-separated). Cases = case rows. `+
    `Revenue = Total Fee (billed). Unique Referrers = distinct doctors/sources. `+
    `<strong>${weekKeys.length} week${weekKeys.length!==1?'s':''}</strong> selected from `+
    `<strong>${s} total</strong> in history. Mode: <strong>${state.mode==='summary'?'Summary (aggregated)':'Weekly Drill-down'}</strong>`;
}

// ── KPI rendering ────────────────────────────────────────────────────────────
function renderKPI(agg,prefix,weekKeys){
  const s=agg.summary;
  const cents=state.centre==='ALL'?allCentres():[state.centre];
  const totalTgt=cents.reduce((acc,c)=>{const t=periodTarget(c,weekKeys);return acc+(t||0);},0);
  function setEl(id,v){const e=document.getElementById(id);if(e)e.textContent=v;}
  setEl(`${prefix}kpi-inv`,fmtN(s.investigations));
  setEl(`${prefix}kpi-cases`,fmtN(s.cases));
  setEl(`${prefix}kpi-refs`,fmtN(s.referrers));
  setEl(`${prefix}kpi-rev`,fmt(s.revenue));
  setEl(`${prefix}kpi-paid`,fmt(s.paid));
  setEl(`${prefix}kpi-due`,fmt(s.due));
  setEl(`${prefix}kpi-rate`,s.revenue>0?(s.paid/s.revenue*100).toFixed(1)+'%':'\u2013');
  const sub=document.getElementById(`${prefix}kpi-rev-sub`);
  if(sub)sub.textContent=totalTgt>0?`${(s.revenue/totalTgt*100).toFixed(1)}% of ${fmt(totalTgt)} target`:'Total fee billed';
}

// ── Summary charts ───────────────────────────────────────────────────────────
function renderTrendChart(agg,weekKeys){
  const labels=agg.trend.map(t=>{
    // Shorten: "Mon 23 – Sun 29 Mar 2026" → "23-29 Mar"
    const w=HISTORY.weeks[t.key];
    if(!w)return t.key;
    const d=t.key.slice(5,7)+'/'+t.key.slice(8,10);
    return w.week_label?w.week_label.replace(/Mon\s/,'').replace(/–\sSun\s/,'-').replace(/\s\d{4}/,''):d;
  });
  const isMulti=labels.length>1;
  document.getElementById('trendTitle').textContent='Revenue by Week (₦)';
  document.getElementById('invTrendTitle').textContent='Investigations by Week';
  mkChart('chartTrend',{type:'bar',data:{labels,datasets:[
    {label:'Revenue',data:agg.trend.map(t=>t.revenue),backgroundColor:'#2E86ABBB',borderColor:'#2E86AB',borderWidth:1},
    {label:'Collected',data:agg.trend.map(t=>t.paid),backgroundColor:'#27AE60BB',borderColor:'#27AE60',borderWidth:1},
  ]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:10}}}},
    scales:{y:{ticks:{callback:v=>'\u20a6'+(v/1e6>=1?(v/1e6).toFixed(1)+'M':(v/1e3).toFixed(0)+'K'),font:{size:9}},grid:{color:'#EEE'}},
      x:{ticks:{font:{size:9}}}}}});
  mkChart('chartInvTrend',{type:'line',data:{labels,datasets:[
    {label:'Investigations',data:agg.trend.map(t=>t.investigations),borderColor:'#2E86AB',backgroundColor:'#2E86AB22',tension:.3,fill:true,pointRadius:3},
    {label:'Cases',data:agg.trend.map(t=>t.cases),borderColor:'#E67E22',backgroundColor:'transparent',tension:.3,borderDash:[5,3],pointRadius:3},
  ]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:10}}}},
    scales:{y:{ticks:{font:{size:9}},grid:{color:'#EEE'}},x:{ticks:{font:{size:9}}}}}});
}

function renderCentreCharts(agg,weekKeys){
  const cents=state.centre==='ALL'?allCentres():[state.centre];
  const activeCents=cents.filter(c=>agg.centreBreakdown.find(r=>r.centre===c));
  const labs=activeCents.map(c=>short(c));
  const revs=activeCents.map(c=>{const r=agg.centreBreakdown.find(x=>x.centre===c);return r?r.revenue:0;});
  const tgts=activeCents.map(c=>periodTarget(c,weekKeys)||0);
  const paids=activeCents.map(c=>{const r=agg.centreBreakdown.find(x=>x.centre===c);return r?r.paid:0;});
  const bgCols=activeCents.map(c=>cCol(c)+'BB');
  mkChart('chartVsTarget',{type:'bar',data:{labels:labs,datasets:[
    {label:'Actual Revenue',data:revs,backgroundColor:bgCols,borderColor:activeCents.map(c=>cCol(c)),borderWidth:1},
    {label:'Period Target',data:tgts,type:'line',borderColor:'#E74C3C',backgroundColor:'transparent',borderWidth:2,borderDash:[6,3],pointRadius:4,pointBackgroundColor:'#E74C3C'},
  ]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:10}}}},
    scales:{y:{ticks:{callback:v=>'\u20a6'+(v/1e6>=1?(v/1e6).toFixed(1)+'M':(v/1e3).toFixed(0)+'K'),font:{size:9}},grid:{color:'#EEE'}},
      x:{ticks:{font:{size:9}}}}}});
  mkChart('chartRevColl',{type:'bar',data:{labels:labs,datasets:[
    {label:'Total Billed',data:revs,backgroundColor:'#2E86ABBB',borderColor:'#2E86AB',borderWidth:1},
    {label:'Collected',data:paids,backgroundColor:'#27AE60BB',borderColor:'#27AE60',borderWidth:1},
  ]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:10}}}},
    scales:{y:{ticks:{callback:v=>'\u20a6'+(v/1e6>=1?(v/1e6).toFixed(1)+'M':(v/1e3).toFixed(0)+'K'),font:{size:9}},grid:{color:'#EEE'}},
      x:{ticks:{font:{size:9}}}}}});
}

// ── Summary tables ───────────────────────────────────────────────────────────
function renderCentreTable(agg,weekKeys){
  const rows=agg.centreBreakdown;
  let totalRev=0,totalPaid=0,totalDue=0,totalInv=0,totalCases=0;
  let html='';
  for(const r of rows){
    const tgt=periodTarget(r.centre,weekKeys);
    const cr=r.revenue>0?(r.paid/r.revenue*100).toFixed(1)+'%':'\u2013';
    const avgFee=r.cases>0?Math.round(r.revenue/r.cases):0;
    totalRev+=r.revenue;totalPaid+=r.paid;totalDue+=r.due;totalInv+=r.investigations;totalCases+=r.cases;
    html+=`<tr><td><span class="dot" style="background:${cCol(r.centre)}"></span>${r.centre}</td>
      <td>${fmtN(r.investigations)}</td><td>${fmtN(r.cases)}</td>
      <td>${fmt(r.revenue)}</td><td>${fmt(r.paid)}</td><td>${fmt(r.due)}</td>
      <td>${fmt(tgt)}</td><td>${achvBadge(r.revenue,tgt)}</td>
      <td>${badge(cr,98,90)}</td><td>${fmt(avgFee)}</td></tr>`;
  }
  const totTgt=allCentres().filter(c=>state.centre==='ALL'||c===state.centre)
    .reduce((s,c)=>{const t=periodTarget(c,weekKeys);return s+(t||0);},0);
  const totCr=totalRev>0?(totalPaid/totalRev*100).toFixed(1)+'%':'\u2013';
  html+=`<tr class="total-row">
    <td>TOTAL</td><td>${fmtN(totalInv)}</td><td>${fmtN(totalCases)}</td>
    <td>${fmt(totalRev)}</td><td>${fmt(totalPaid)}</td><td>${fmt(totalDue)}</td>
    <td>${fmt(totTgt||null)}</td><td>${achvBadge(totalRev,totTgt)}</td>
    <td>${totCr}</td><td>\u2013</td></tr>`;
  document.getElementById('centreTbody').innerHTML=html;
}
function renderRefTable(refs,prefix){
  const centreRevMap={};
  if(!prefix){
    // build centre total rev map for % calc
    document.getElementById('refTbody'+(prefix||'')).innerHTML=
      refs.map((r,i)=>`<tr><td>${i+1}</td><td><span class="dot" style="background:${cCol(r.centre)}"></span>${r.centre}</td>
        <td>${r.ref}</td><td>${fmtN(r.cases)}</td><td>${fmt(r.revenue)}</td><td>\u2013</td></tr>`).join('');
  }
}
function renderRefTable2(refs,centreBreakdown){
  const revMap={};
  centreBreakdown.forEach(r=>{revMap[r.centre]=r.revenue;});
  document.getElementById('refTbody').innerHTML=refs.map((r,i)=>{
    const pct=revMap[r.centre]>0?(r.revenue/revMap[r.centre]*100).toFixed(1)+'%':'\u2013';
    return`<tr><td>${i+1}</td><td><span class="dot" style="background:${cCol(r.centre)}"></span>${r.centre}</td>
      <td>${r.ref}</td><td>${fmtN(r.cases)}</td><td>${fmt(r.revenue)}</td><td>${pct}</td></tr>`;
  }).join('');
}
function renderAgentTable(agents,tbodyId){
  document.getElementById(tbodyId).innerHTML=agents.map((a,i)=>{
    const cr=a.revenue>0?(a.paid/a.revenue*100).toFixed(1)+'%':'\u2013';
    return`<tr><td>${i+1}</td><td>${a.agent}</td><td>${fmtN(a.investigations)}</td>
      <td>${fmtN(a.cases)}</td><td>${fmt(a.revenue)}</td><td>${fmt(a.paid)}</td>
      <td>${badge(cr,98,90)}</td></tr>`;
  }).join('');
}

// ── Drill-down rendering (single week) ───────────────────────────────────────
function renderDrill(weekKey){
  const w=HISTORY.weeks[weekKey];
  if(!w){document.getElementById('drillNote').textContent='No data for selected week.';return;}
  const fcentre=state.centre;

  document.getElementById('drillNote').innerHTML=
    `<strong>${w.week_label}</strong> \u2014 ${w.quarter||''} targets applied. `+
    `Cancelled: ${w.summary.cancelled||0}. Tests = investigations; Cases = case rows.`;

  // KPIs
  const daily=fcentre==='ALL'?w.daily:w.daily.filter(r=>r['Collection centre']===fcentre);
  const weekly=fcentre==='ALL'?w.weekly:w.weekly.filter(r=>r['Collection centre']===fcentre);
  let dRev=0,dPaid=0,dDue=0,dInv=0,dCases=0,dPat=0,dRefCnt=0;
  daily.forEach(r=>{dRev+=(r.revenue||0);dPaid+=(r.paid||0);dDue+=(r.due||0);dInv+=(r.investigations||0);dCases+=(r.cases||0);});
  weekly.forEach(r=>{dPat+=(r.patients||0);dRefCnt+=(r.unique_referrers||0);});
  ['inv','cases','refs','rev','paid','due','rate'].forEach(k=>{
    const el=document.getElementById('dkpi-'+k);if(el)el.textContent='\u2013';
  });
  document.getElementById('dkpi-inv').textContent=fmtN(dInv);
  document.getElementById('dkpi-cases').textContent=fmtN(dCases);
  document.getElementById('dkpi-refs').textContent=fmtN(dRefCnt);
  document.getElementById('dkpi-rev').textContent=fmt(dRev);
  document.getElementById('dkpi-paid').textContent=fmt(dPaid);
  document.getElementById('dkpi-due').textContent=fmt(dDue);
  document.getElementById('dkpi-rate').textContent=dRev>0?(dPaid/dRev*100).toFixed(1)+'%':'\u2013';
  const dTgts=(fcentre==='ALL'?allCentres():[fcentre]).reduce((s,c)=>{const t=periodTarget(c,[weekKey]);return s+(t||0);},0);
  const drSub=document.getElementById('dkpi-rev-sub');
  if(drSub)drSub.textContent=dTgts>0?`${(dRev/dTgts*100).toFixed(1)}% of ${fmt(dTgts)} weekly target`:'vs weekly target';

  // Daily charts
  const dates=w.dates||[...new Set(daily.map(r=>r['Date']))].sort();
  const dateLabels=w.date_labels||(()=>{
    const d=new Date(dates[0]);
    return dates.map(ds=>{const dt=new Date(ds);return dt.toLocaleDateString('en-GB',{weekday:'short',day:'numeric',month:'short'});});
  })();
  const cents=fcentre==='ALL'?(w.centres||[...new Set(daily.map(r=>r['Collection centre']))].sort()):[fcentre];

  mkChart('dChartDailyInv',{type:'bar',data:{labels:dateLabels,datasets:cents.map((c,i)=>({
    label:short(c),
    data:dates.map(d=>{const r=daily.find(x=>x['Date']===d&&x['Collection centre']===c);return r?r.investigations:0;}),
    backgroundColor:COLOURS[i%COLOURS.length]+'BB',borderColor:COLOURS[i%COLOURS.length],borderWidth:1,
  }))},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:9}}}},
    scales:{x:{stacked:true,ticks:{font:{size:9}}},y:{stacked:true,ticks:{font:{size:9}},grid:{color:'#EEE'}}}}});

  mkChart('dChartDailyRev',{type:'bar',data:{labels:dateLabels,datasets:cents.map((c,i)=>({
    label:short(c),
    data:dates.map(d=>{const r=daily.find(x=>x['Date']===d&&x['Collection centre']===c);return r?r.revenue:0;}),
    backgroundColor:COLOURS[i%COLOURS.length]+'BB',borderColor:COLOURS[i%COLOURS.length],borderWidth:1,
  }))},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:9}}}},
    scales:{x:{stacked:true,ticks:{font:{size:9}}},y:{stacked:true,ticks:{callback:v=>'\u20a6'+(v/1e3).toFixed(0)+'K',font:{size:9}},grid:{color:'#EEE'}}}}});

  // vs Target chart
  const wRevByCentre=cents.map(c=>{const r=weekly.find(x=>x['Collection centre']===c);return r?r.revenue:0;});
  const wTgtByCentre=cents.map(c=>periodTarget(c,[weekKey])||0);
  mkChart('dChartVsTarget',{type:'bar',data:{labels:cents.map(c=>short(c)),datasets:[
    {label:'Actual',data:wRevByCentre,backgroundColor:cents.map((c,i)=>COLOURS[i%COLOURS.length]+'BB'),borderColor:cents.map((c,i)=>COLOURS[i%COLOURS.length]),borderWidth:1},
    {label:'Target',data:wTgtByCentre,type:'line',borderColor:'#E74C3C',backgroundColor:'transparent',borderWidth:2,borderDash:[6,3],pointRadius:4,pointBackgroundColor:'#E74C3C'},
  ]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:9}}}},
    scales:{y:{ticks:{callback:v=>'\u20a6'+(v/1e3).toFixed(0)+'K',font:{size:9}},grid:{color:'#EEE'}},x:{ticks:{font:{size:9}}}}}});

  // Daily collection rate
  mkChart('dChartCollRate',{type:'line',data:{labels:dateLabels,datasets:cents.map((c,i)=>({
    label:short(c),
    data:dates.map(d=>{const r=daily.find(x=>x['Date']===d&&x['Collection centre']===c);return(r&&r.revenue)?+(r.paid/r.revenue*100).toFixed(1):null;}),
    borderColor:COLOURS[i%COLOURS.length],backgroundColor:'transparent',tension:.3,pointRadius:3,
  }))},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{font:{size:9}}}},
    scales:{y:{min:0,max:100,ticks:{callback:v=>v+'%',font:{size:9}},grid:{color:'#EEE'}},x:{ticks:{font:{size:9}}}}}});

  // Daily table
  const sortedDaily=[...daily].sort((a,b)=>a['Date']===b['Date']?a['Collection centre'].localeCompare(b['Collection centre']):a['Date'].localeCompare(b['Date']));
  document.getElementById('dailyTbody').innerHTML=sortedDaily.map(r=>{
    const c=r['Collection centre'];
    const dt=(w.centre_targets||{})[c]?.daily;
    const rev=r.revenue||0;const cr=rev>0?(r.paid/rev*100).toFixed(1)+'%':'\u2013';
    return`<tr><td>${r['Date']}</td>
      <td><span class="dot" style="background:${cCol(c)}"></span>${c}</td>
      <td>${fmtN(r.investigations)}</td><td>${fmtN(r.cases)}</td><td>${fmtN(r.patients)}</td>
      <td>${fmtN(r.unique_referrers)}</td><td>${fmt(rev)}</td><td>${fmt(r.paid)}</td><td>${fmt(r.due)}</td>
      <td>${fmt(dt)}</td><td>${achvBadge(rev,dt)}</td><td>${badge(cr,98,90)}</td></tr>`;
  }).join('');

  // Weekly summary table
  document.getElementById('weeklyTbody').innerHTML=weekly.map(r=>{
    const c=r['Collection centre'];
    const wt=periodTarget(c,[weekKey]);const rev=r.revenue||0;
    const cr=rev>0?(r.paid/rev*100).toFixed(1)+'%':'\u2013';
    return`<tr><td><span class="dot" style="background:${cCol(c)}"></span>${c}</td>
      <td>${fmtN(r.investigations)}</td><td>${fmtN(r.cases)}</td><td>${fmtN(r.patients)}</td>
      <td>${fmtN(r.unique_referrers)}</td><td>${fmt(rev)}</td><td>${fmt(r.paid)}</td><td>${fmt(r.due)}</td>
      <td>${fmt(wt)}</td><td>${achvBadge(rev,wt)}</td><td>${badge(cr,98,90)}</td></tr>`;
  }).join('');

  // Drill referrers + agents
  const dRefs=(fcentre==='ALL'?w.top_referrers:(w.top_referrers||[]).filter(r=>r['Collection centre']===fcentre))||[];
  document.getElementById('dRefTbody').innerHTML=dRefs.slice(0,30).map((r,i)=>
    `<tr><td>${i+1}</td><td><span class="dot" style="background:${cCol(r['Collection centre'])}"></span>${r['Collection centre']}</td>
    <td>${r['Referrer']}</td><td>${fmtN(r.cases)}</td><td>${fmt(r.revenue)}</td></tr>`).join('');
  renderAgentTable(w.agent_total||[],'dAgentTbody');
}

// ── Master render ─────────────────────────────────────────────────────────────
function renderAll(){
  const weekKeys=currentWeekKeys();
  updatePeriodLabel(weekKeys);
  updateNote(weekKeys);
  document.getElementById('weekCount').textContent=`${weekKeys.length} week${weekKeys.length!==1?'s':''} loaded`;

  if(state.mode==='summary'){
    const agg=aggregate(weekKeys,state.centre);
    renderKPI(agg,'',weekKeys);
    renderTrendChart(agg,weekKeys);
    renderCentreCharts(agg,weekKeys);
    renderCentreTable(agg,weekKeys);
    renderRefTable2(agg.topReferrers,agg.centreBreakdown);
    renderAgentTable(agg.agents,'agentTbody');
  } else {
    // Drill-down
    const wk=state.week||(weekKeys.length>0?weekKeys[weekKeys.length-1]:null);
    if(wk){
      // Update week select to match
      document.getElementById('wkSel').value=wk;
      renderDrill(wk);
    } else {
      document.getElementById('drillNote').textContent='Please select a specific week to view the drill-down.';
    }
  }
}

// ── Initialization ────────────────────────────────────────────────────────────
function init(){
  const years=getYears();
  state.year=years[0];
  populateYears(years);
  const months=getMonths(state.year);
  state.month=months[months.length-1]||0; // default to most recent month
  document.getElementById('monSel').value=state.month;
  populateMonths(state.year);
  document.getElementById('monSel').value=state.month;
  const weekKeys=getWeekKeys(state.year,state.month);
  state.week=weekKeys.length>0?weekKeys[weekKeys.length-1]:null; // default to most recent week
  populateWeeks(state.year,state.month);
  document.getElementById('wkSel').value=state.week||'';

  // Centre tabs
  const bar=document.getElementById('centreBar');
  allCentres().forEach((c,i)=>{
    const d=document.createElement('div');d.className='tab';
    d.innerHTML=`<span class="dot" style="background:${COLOURS[i%COLOURS.length]}"></span>${short(c)}`;
    d.onclick=()=>filterCentre(c,d);bar.appendChild(d);
  });

  document.getElementById('centreList').textContent=allCentres().join('  |  ');
  renderAll();
}
document.addEventListener('DOMContentLoaded',()=>init());
</script></body></html>
""")

    content = ''.join(p)
    html_path = os.path.join(WORKSPACE, 'Sure_Weekly_Dashboard_Multi.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"Multi-period HTML saved: {html_path}")
    return html_path

if __name__ == '__main__':
    main()
